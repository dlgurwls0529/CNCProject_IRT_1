# -*- coding: utf-8 -*-
import math
import numpy as np
import pandas as pd
import girth
import os
import json
import copy
import pyexpat
from openpyxl import load_workbook
from openpyxl import Workbook
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler
from factor_analyzer import FactorAnalyzer
from factor_analyzer.factor_analyzer import calculate_kmo
import matplotlib.pyplot as plt

#https://coding-kindergarten.tistory.com/75

class Test:
    # init 안쓰고 eager init 하면 리플렉션 인식 안됨
    def __init__(self,
                 no=None,
                 name=None,
                 birth=None,
                 phone_number=None,
                 teacher=None,
                 level=None,
                 class_number=None,
                 sector=None,
                 school=None,
                 grade=None,
                 lc=None,
                 gr=None,
                 rc=None
                 ):
        self.no = no
        self.name = name
        self.birth = birth
        self.phone_number = phone_number
        self.teacher = teacher
        self.level = level
        self.class_number = class_number
        self.sector = sector
        self.school = school
        self.grade = grade
        self.lc = []
        self.gr = []
        self.rc = []

class DefaultCheckedTest(Test):
    def __init__(self):
        super().__init__()
        self.errata = []

# 시험 점수 최종 채점된 데이터 저장
class WeightCheckedTest(Test):
    def __init__(self, score):
        super().__init__()
        self.score = score

class FinalTestInfo:
    def __init__(self,
                 no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                 lc, gr, rc,
                 score, rank_total, rank_level, stanine, score_area, rank_total_area, rank_level_area, stanine_area):

        self.no = no
        self.name = name
        self.birth = birth
        self.phone_number = phone_number
        self.teacher = teacher
        self.level = level
        self.class_number = class_number
        self.sector = sector
        self.school = school
        self.grade = grade
        self.lc = []
        self.gr = []
        self.rc = []
        self.score = score
        self.rank_total = rank_total
        self.rank_level = rank_level
        self.stanine = stanine
        self.score_area = score_area
        self.rank_total_area = rank_total_area
        self.rank_level_area = rank_level_area
        self.stanine_area = stanine_area

class Answer:
    def __init__(self,
                 domain: str,
                 item_number: str,
                 item_answer: str,
                 category: str,
                 source: str
                 ) -> object:
        self.domain = self.get_domain(int(str(item_number).replace('번', '')))
        self.item_number = int(str(item_number).replace('번', ''))
        self.item_answer = item_answer
        self.category = category
        self.source = source

    @staticmethod
    def get_domain(item_number):
        if item_number is None:
            raise Exception('item number blank')

        else:
            if 1 <= item_number <= 30:
                return 'LC'

            elif 31 <= item_number <= 45:
                return 'GR'

            elif 46 <= item_number <= 85:
                return 'RC'

            else:
                raise Exception('invalid item number')

# 배점 할당된 답
class WeightedAnswer(Answer):
    def __init__(self, domain: str, item_number: str, item_answer: str, category: str, source: str, point):
        super().__init__(domain, item_number, item_answer, category, source)
        self.point = point

def is_column_attribute(column):
    if 1 <= column <= 10:
        return True
    else:
        return False

def is_column_lc(column):
    if 11 <= column <= 40:
        return True
    else:
        return False

def is_column_gr(column):
    if 41 <= column <= 55:
        return True
    else:
        return False

def is_column_rc(column):
    if 56 <= column <= 95:
        return True
    else:
        return False

def level2len(test_list, level):
    if 1 <= int(level) <= 7:
        return

def read_answer_excel(path, sheet):
    answer_list = []

    wb = load_workbook(path)
    ws = wb[sheet]

    row_begin = 2
    row_end = 86

    for i in range(row_begin, row_end + 1):
        answer = Answer(
            ws.cell(i, 1).value,
            ws.cell(i, 2).value,
            ws.cell(i, 3).value,
            ws.cell(i, 4).value,
            ws.cell(i, 5).value
        )
        answer_list.append(answer)

    return answer_list

def read_test_excel(path, sheet):
    test_list = []

    workbook = load_workbook(path)
    worksheet = workbook[sheet]

    row_begin = 3
    row_end = 12

    column_begin = 1
    column_end = 95

    # row iteration
    for i in range(row_begin, row_end + 1):
        test = Test()

        field_count = 0
        for j in range(column_begin, column_end + 1):
            cell = worksheet.cell(i, j).value

            if is_column_attribute(j):
                setattr(test, str(list(vars(test))[field_count]), cell)
                field_count = field_count + 1

            elif is_column_lc(j):
                # swallow copy
                test.lc.append(cell)

            elif is_column_gr(j):
                test.gr.append(cell)

            elif is_column_rc(j):
                test.rc.append(cell)

            else:
                raise IndexError('list index out of range')

        test_list.append(test)

    return test_list

def get_answer_list_by_level(level, answer_1to7_list, answer_8to11_list):
    if 1 <= level <= 7:
        return answer_1to7_list

    elif 8 <= level <= 11:
        return answer_8to11_list

    else:
        raise Exception('invalid level')

def get_errata_list(test_list, answer_1to7_list, answer_8to11_list):
    errata_list = []

    for i in test_list:
        answer_list = get_answer_list_by_level(int(i.level), answer_1to7_list, answer_8to11_list)
        current_test = i.lc + i.gr + i.rc
        current_errata = []

        if len(answer_list) != len(current_test):
            raise Exception('answer list size is not equal to test')

        else:
            for j in range(len(answer_list)):
                if str(current_test[j]) == str(answer_list[j].item_answer):
                    current_errata.append(1)
                else:
                    current_errata.append(0)

        errata_list.append(current_errata)

    return np.transpose(np.array(errata_list))

def get_default_checked_test_list(test_list, answer_1to7_list, answer_8to11_list):
    default_checked_test_list = []

    for i in test_list:
        answer_list = get_answer_list_by_level(int(i.level), answer_1to7_list, answer_8to11_list)
        current_test = i.lc + i.gr + i.rc
        current_default_checked_test = DefaultCheckedTest()

        for j in vars(i):
            setattr(current_default_checked_test, j, getattr(i, j))

        if len(answer_list) != len(current_test):
            raise Exception('answer list size is not equal to test')

        else:
            for j in range(len(answer_list)):
                if str(current_test[j]) == str(answer_list[j].item_answer):
                    current_default_checked_test.errata.append(1)
                else:
                    current_default_checked_test.errata.append(0)

        default_checked_test_list.append(current_default_checked_test)

    return default_checked_test_list

def my_ceil(num, digit):
    num = np.ceil(num * (10 ** digit)) / (10 ** digit)
    return num

# return weighted answer 1 to 7 and 8 to 11
def get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list):
    default_check_test_list = get_default_checked_test_list(test_list, answer_1to7_list, answer_8to11_list)

    errata_list_level_1to7 = [x.errata for x in default_check_test_list if 1 <= int(x.level) <= 7]
    errata_list_level_8to11 = [x.errata for x in default_check_test_list if 8 <= int(x.level) <= 11]

    estimate_1to7 = np.array(girth.twopl_mml(np.transpose(np.array(errata_list_level_1to7)))['Difficulty'])
    estimate_8to11 = np.array(girth.twopl_mml(np.transpose(np.array(errata_list_level_8to11)))['Difficulty'])

    weight_list_1to7 = np.zeros(85)
    weight_list_8to11 = np.zeros(85)

    for i in range(len(estimate_1to7)):
        index = estimate_1to7.argmax()
        estimate_1to7[index] = -np.inf

        if 0 <= i <= 16:
            weight_list_1to7[index] = 1.6666666635
        elif 17 <= i <= 33:
            weight_list_1to7[index] = 1.3333333325
        elif 34 <= i <= 50:
            weight_list_1to7[index] = 1.0000000015
        elif 51 <= i <= 67:
            weight_list_1to7[index] = 0.6666666705
        elif 68 <= i <= 84:
            weight_list_1to7[index] = 0.3333333310
        else:
            raise Exception()

    for i in range(len(estimate_8to11)):
        index = estimate_8to11.argmax()
        estimate_8to11[index] = -np.inf

        if 0 <= i <= 16:
            weight_list_8to11[index] = 1.6666666635
        elif 17 <= i <= 33:
            weight_list_8to11[index] = 1.3333333325
        elif 34 <= i <= 50:
            weight_list_8to11[index] = 1.0000000015
        elif 51 <= i <= 67:
            weight_list_8to11[index] = 0.6666666705
        elif 68 <= i <= 84:
            weight_list_8to11[index] = 0.3333333310
        else:
            raise Exception()

    weighted_answer_list_1to7 = [WeightedAnswer(
        answer_1to7_list[i].domain,
        answer_1to7_list[i].item_number,
        answer_1to7_list[i].item_answer,
        answer_1to7_list[i].category,
        answer_1to7_list[i].source,
        list(weight_list_1to7)[i]
    ) for i in range(len(answer_1to7_list))]

    weighted_answer_list_8to11 = [WeightedAnswer(
        answer_8to11_list[i].domain,
        answer_8to11_list[i].item_number,
        answer_8to11_list[i].item_answer,
        answer_8to11_list[i].category,
        answer_8to11_list[i].source,
        list(weight_list_8to11)[i]
    ) for i in range(len(answer_8to11_list))]

    """
    mms = MinMaxScaler()

    estimates_1to7 = np.array(girth.twopl_mml(np.transpose(np.array(
        [x.errata for x in default_check_test_list_level_1to7]
    )))['Difficulty']).reshape(-1, 1)

    mms.fit(estimates_1to7)
    estimates_1to7 = mms.transform(estimates_1to7)
    estimates_1to7 = (85/np.sum(estimates_1to7))*estimates_1to7

    estimates_8to11 = np.array(girth.twopl_mml(np.transpose(np.array(
        [x.errata for x in default_check_test_list_level_8to11]
    )))['Difficulty']).reshape(-1, 1)

    # 85/(a + b + c) = k

    mms.fit(estimates_8to11)
    estimates_8to11 = mms.transform(estimates_8to11)
    estimates_8to11 = (85/np.sum(estimates_8to11))*estimates_8to11

    weighted_answer_list_1to7 = [WeightedAnswer(
        answer_1to7_list[i].domain,
        answer_1to7_list[i].item_number,
        answer_1to7_list[i].item_answer,
        answer_1to7_list[i].category,
        answer_1to7_list[i].source,
        list(np.array(estimates_1to7).flatten())[i]
    ) for i in range(len(answer_1to7_list))]

    weighted_answer_list_8to11 = [WeightedAnswer(
        answer_8to11_list[i].domain,
        answer_8to11_list[i].item_number,
        answer_8to11_list[i].item_answer,
        answer_8to11_list[i].category,
        answer_8to11_list[i].source,
        list(np.array(estimates_8to11).flatten())[i]
    ) for i in range(len(answer_8to11_list))]
    """

    return weighted_answer_list_1to7, weighted_answer_list_8to11

def get_weight_checked_test_list(default_checked_test_list, weighted_answer_1to7_list, weighted_answer_8to11_list):
    weight_check_test_list = []

    for i in range(len(default_checked_test_list)):
        if 1 <= int(default_checked_test_list[i].level) <= 7:
            score = np.dot(np.array(default_checked_test_list[i].errata), np.array([
                x.point for x in weighted_answer_1to7_list
            ]))

        elif 8 <= int(default_checked_test_list[i].level) <= 11:
            score = np.dot(np.array(default_checked_test_list[i].errata), np.array([
                x.point for x in weighted_answer_8to11_list
            ]))

        else:
            raise Exception('invalid value : level')

        currentWeightCheckTest = WeightCheckedTest(float(score))

        for j in vars(Test()):
            setattr(currentWeightCheckTest, j, getattr(default_checked_test_list[i], j))

        weight_check_test_list.append(currentWeightCheckTest)

    for i in range(len(weight_check_test_list)):
        if 1 <= int(weight_check_test_list[i].level) <= 7:
            weight_check_test_list[i].score = np.round(180.0 + (420.0 *
                float(weight_check_test_list[i].score)) / 85.0
            )
        elif 8 <= int(weight_check_test_list[i].level) <= 11:
            weight_check_test_list[i].score = np.round(240.0 + (360.0 *
                float(weight_check_test_list[i].score)) / 85.0
            )
        else:
            raise Exception()


    return weight_check_test_list

def read_pseudo_test_excel(path, sheet):
    wb = load_workbook(path)
    ws = wb[sheet]

    row_begin = 3
    row_end = 4369

    column_begin = 1
    column_end = 95

    test_list = []

    for i in range(row_begin, row_end + 1):
        current_test = Test()

        is_missing_row = False
        for j in range(column_begin, column_end + 1):
            cell = ws.cell(i, j).value

            if is_column_attribute(j):
                if j == 3 and cell == '미응시':
                    is_missing_row = True
                    break
                elif j == 6:
                    if str(cell) == '0.1':
                        cell = '1'
                    elif str(cell) == 'har':
                        cell = '2'
                    elif str(cell) == 'stan':
                        cell = '3'
                    elif str(cell) == 'yale':
                        cell = '4'
                    elif str(cell) == 'prin':
                        cell = '5'
                    elif str(cell) == 'col':
                        cell = '6'
                    elif (str(cell) == 'duke') or (str(cell) == 'Duke'):
                        cell = '7'
                    elif str(cell) == 'P-B':
                        cell = '8'
                    elif str(cell) == 'P-A':
                        cell = '9'
                    elif str(cell) == 'Advan':
                        cell = '10'
                    elif str(cell) == 'Inter':
                        cell = '11'
                    else:
                        raise Exception('invalid level')

                setattr(current_test, str(list(vars(current_test))[j - 1]), cell)

            elif is_column_lc(j):
                # swallow copy
                current_test.lc.append(cell)

            elif is_column_gr(j):
                current_test.gr.append(cell)

            elif is_column_rc(j):
                current_test.rc.append(cell)

            else:
                raise IndexError('list index out of range')

        if is_missing_row is not True:
            test_list.append(current_test)

    return test_list

def rank2stanine(rank:float):
    stanine = -1

    if 0.0 < rank <= 4.0:
        stanine = 1
    elif 4.0 < rank <= 11.0:
        stanine = 2
    elif 11.0 < rank <= 23.0:
        stanine = 3
    elif 23.0 < rank <= 40.0:
        stanine = 4
    elif 40.0 < rank <= 60.0:
        stanine = 5
    elif 60.0 < rank <= 77.0:
        stanine = 6
    elif 77.0 < rank <= 89.0:
        stanine = 7
    elif 89.0 < rank <= 96.0:
        stanine = 8
    elif 96.0 < rank <= 100.0:
        stanine = 9
    else:
        raise Exception()

    return stanine

def score2rank(score_list:[]):
    rank = [1]

    # [2.0, 2.0, 9.0, 4.0, 3.0, 4.0, 3.0, 4.0]

    sorted_score_list = np.flip(np.sort(np.array(score_list)))
    # [9.0, 4.0, 4.0, 4.0, 3.0, 3.0, 2.0, 2.0]
    #  i:         1    2    3    4    5    6    7
    # rc:    1    2    2    2    5    5    7    7
    # ec:    0    0    1    2    0    1    0    1
    #  r:    1    2    2    2    5    5    7    7
    rank_count = 1
    current_equal_count = 0

    for i in range(1, len(sorted_score_list)):
        if str(sorted_score_list[i - 1]) == str(sorted_score_list[i]):
            current_equal_count = current_equal_count + 1
        else:
            rank_count = rank_count + 1 + current_equal_count
            current_equal_count = 0
        rank.append(rank_count)

    result = np.zeros(len(rank))
    for i in range(len(score_list)):
        for j in range(len(sorted_score_list)):
            if str(score_list[i]) == str(sorted_score_list[j]):
                result[i] = int(rank[j])
                break




    return result

def get_tetrachoric_correlation(response_matrix):
    df = pd.DataFrame(
        response_matrix,
        columns=[str('Item' + str(x)) for x in range(1, 86)]
    )

    return df.corr()

def get_score_list_group_level(test_list, score_list):
    # level is fixed
    index_list_level = [[i for i in range(len(test_list)) if test_list[i].level == str(j)]
                        for j in range(1, 12)]

    ranking_list_ordered_by_level_score = np.zeros(len(score_list))
    index_list_level_index_list_score = []


    for i in range(11):
        current_level_index = np.array([int(x) for x in index_list_level[i]])

        if len(current_level_index) == 0:
            index_list_level_index_list_score.append([])
        else:
            current_level_index_score = score2rank(
                np.array([float(x) for x in score_list])[np.array(current_level_index)]
            )
            index_list_level_index_list_score.append(current_level_index_score)

    for i in range(11):
        for j in range(len(index_list_level[i])):
            ranking_list_ordered_by_level_score[index_list_level[i][j]] = \
                index_list_level_index_list_score[i][j] * 100 / len(index_list_level_index_list_score[i])

    return ranking_list_ordered_by_level_score

class UnitTest:

    # noinspection PyMethodMayBeStatic
    def read_test_excel_test(self):
        path = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet = 'Data'
        test_list = read_test_excel(path, sheet)

        for i in test_list:
            for j in vars(i):
                print(j, ' : ', getattr(i, j))
            print(' ')

    # noinspection PyMethodMayBeStatic
    def sample_data_test(self):
        np.set_printoptions(threshold=np.inf, linewidth=np.inf)
        df = pd.read_csv("https://github.com/jmnote/zdata/raw/master/github.com/cran/ltm/data/LSAT.csv")
        m = np.transpose(df.values)
        print(m)

    # noinspection PyMethodMayBeStatic
    def irt_test(self):
        df = pd.read_csv("https://github.com/jmnote/zdata/raw/master/github.com/cran/ltm/data/LSAT.csv")
        m = np.transpose(df.values)
        estimates = girth.threepl_mml(m)
        pd.DataFrame(estimates)
        print(estimates)

        """ 3PL IRT 모델에서 파라미터를 추정한다.
            
        인자:
            dataset: [items x participants] matrix of True/False Values
            options: dictionary with updates to default options

        Returns:
            discrimination: (1d array) estimate of item discriminations
            difficulty: (1d array) estimates of item diffiulties
            guessing: (1d array) estimates of item guessing

        Options:
            * max_iteration: int
            * distribution: callable
            * quadrature_bounds: (float, float)
            * quadrature_n: int
        """

    # noinspection PyMethodMayBeStatic
    def test_field_injection_test(self):
        path = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet = 'Data'

        workbook = load_workbook(path)
        worksheet = workbook[sheet]

        row_begin = 3
        row_end = 12

        column_begin = 1
        column_end = 95

        test = Test()

        field_count = 0
        for i in range(column_begin, column_end+1):
            cell = worksheet.cell(row_begin, i).value

            if is_column_attribute(i):
                setattr(test, str(list(vars(test))[field_count]), cell)

            elif is_column_lc(i):
                # swallow copy
                test.lc.append(cell)

            elif is_column_gr(i):
                test.gc.append(cell)

            elif is_column_rc(i):
                test.rc.append(cell)

            else:
                raise IndexError('list index out of range')

            field_count = field_count + 1

        for i in vars(test):
            print(i, ' : ', getattr(test, i))

    # noinspection PyMethodMayBeStatic
    def answer_1to7_field_injection_test(self):
        path = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet = '정답'

        wb = load_workbook(path)
        ws = wb[sheet]

        row_sample = 6

        answer = Answer(
            ws.cell(row_sample, 1).value,
            ws.cell(row_sample, 2).value,
            ws.cell(row_sample, 3).value,
            ws.cell(row_sample, 4).value,
            ws.cell(row_sample, 5).value
        )

        for i in vars(answer):
            print(i, ' : ', getattr(answer, i))

    # noinspection PyMethodMayBeStatic
    def excel_access_test(self):
        path = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet = 'Data'

        workbook = load_workbook(path)
        worksheet = workbook[sheet]

        for i in range(1, 10):
            print(worksheet.cell(1, i).value)

    # noinspection PyMethodMayBeStatic
    def test_reflection_test(self):
        test = Test()
        test.name = 'test_name'
        for i in vars(test):
            print(i, ' : ', getattr(test, i))

    # noinspection PyMethodMayBeStatic
    def read_answer_1to7_excel_test(self):
        path = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet = '정답'

        answer_list = read_answer_excel(path, sheet)

        for i in answer_list:
            for j in vars(i):
                print(j, ' : ', getattr(i, j))
            print('')

    # noinspection PyMethodMayBeStatic
    def read_answer_8to11_excel_test(self):
        pass

    # noinspection PyMethodMayBeStatic
    def get_answer_list_by_level_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        for i in test_list:
            temp = '[ '

            for j in get_answer_list_by_level(int(i.level), answer_1to7_list, answer_8to11_list):
                temp = temp + str(j.item_answer) + ' '

            temp = temp + ']'

            print(temp)

    # noinspection PyMethodMayBeStatic
    def errata_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        errata_list = get_errata_list(test_list, answer_1to7_list, answer_8to11_list)

        np.set_printoptions(threshold=np.inf, linewidth=np.inf)
        print(errata_list)

    # noinspection PyMethodMayBeStatic
    def point_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        default_check_test_list = get_default_checked_test_list(test_list, answer_1to7_list, answer_8to11_list)

        errata_list = []

        for i in default_check_test_list:
            errata_list.append(i.errata)

        transposed_errata = np.transpose(np.array(errata_list))

        estimates_one = girth.onepl_mml(transposed_errata)
        print(estimates_one['Difficulty'])

        # 결과는 item by test 행렬

    # noinspection PyMethodMayBeStatic
    def get_weighted_answer_list_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        print([x.point for x in weighted_answer_1to7_list])
        print([x.point for x in weighted_answer_8to11_list])

    # noinspection PyMethodMayBeStatic
    def get_default_checked_test_list_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        for i in default_checked_test_list:
            print(i.errata)

    # noinspection PyMethodMayBeStatic
    def get_weight_checked_test_list_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )

        for i in vars(Test()):
            field_name = i
            print(field_name, ' : ', getattr(weight_check_test_list[0], field_name))

        print('errata : ', default_checked_test_list[0].errata)
        print('point : ', [x.point for x in weighted_answer_1to7_list])
        print('default_score : ', np.sum(np.array(default_checked_test_list[0].errata)))
        print('weight_score : ', weight_check_test_list[0].score)

    # noinspection PyMethodMayBeStatic
    def default_checked_test_reflection_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        for i in vars(Test()):
            field_name = i
            print(field_name, ' : ', getattr(default_checked_test_list[3], field_name))

    # noinspection PyMethodMayBeStatic
    def equality_test(self):
        print('1' != '1,2,2')
        print('1' == '')
        print('1' == '1,2')

    # noinspection PyMethodMayBeStatic
    def read_pseudo_test_excel_test(self):
        path = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet = 'Data'

        test_list = read_pseudo_test_excel(path, sheet)
        print(len(test_list))
        for i in test_list[4360:]:
            for j in vars(i):
                print(j, ' : ', getattr(i, j))
            print(' ')

    # noinspection PyMethodMayBeStatic
    def get_pseudo_weight_checked_test_list_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )


        """
        write_wb = Workbook()
        write_ws = write_wb.active

        for i in range(len(weight_check_test_list)):
            column_count = 1
            for j in vars(Test()):
                if is_column_attribute(column_count):
                    write_ws.cell(i+1, column_count, getattr(weight_check_test_list[i], j))
                    column_count = column_count + 1
                else:
                    break

            write_ws.cell(i+1, column_count, np.sum(np.array(default_checked_test_list[i].errata)))
            column_count = column_count + 1

            write_ws.cell(i+1, column_count, weight_check_test_list[i].score)
            column_count = column_count + 1

            for j in range(len(default_checked_test_list[i].errata)):
                write_ws.cell(i+1, column_count, default_checked_test_list[i].errata[j])
                column_count = column_count + 1

            if 1 <= int(weight_check_test_list[i].level) <= 7:
                point_list = [x.point for x in weighted_answer_1to7_list]

            elif 8 <= int(weight_check_test_list[i].level) <= 11:
                point_list = [x.point for x in weighted_answer_8to11_list]

            else:
                raise Exception('invalid level')

            for j in range(len(point_list)):
                write_ws.cell(i+1, column_count, round(point_list[j], 2))
                column_count = column_count + 1

        write_wb.save('/Users/dlgur/OneDrive/문서/Reports/결과파일.xlsx')
        """

        for i in range(10):
            for j in vars(Test()):
                field_name = j
                print(field_name, ' : ', getattr(weight_check_test_list[i], field_name))

            print('errata : ', default_checked_test_list[i].errata)

            if 1 <= int(weight_check_test_list[i].level) <= 7:
                print('point : ', np.round([x.point for x in weighted_answer_1to7_list], 2))

            elif 8 <= int(weight_check_test_list[i].level) <= 11:
                print('point : ', np.round([x.point for x in weighted_answer_8to11_list], 2))

            else:
                raise Exception('invalid level')

            print('default_score : ', np.sum(np.array(default_checked_test_list[i].errata)))
            print('weight_score : ', weight_check_test_list[i].score)

            print('')

    # noinspection PyMethodMayBeStatic
    def get_pseudo_default_checked_test_list_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx"
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        for i in default_checked_test_list:
            print(i.errata)

    # noinspection PyMethodMayBeStatic
    def get_tetrachoric_correlation_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx"
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        response_matrix = [x.errata for x in default_checked_test_list]

        print(get_tetrachoric_correlation(response_matrix))

    # noinspection PyMethodMayBeStatic
    def get_kmo_validation_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx"
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        response_matrix = [x.errata for x in default_checked_test_list]

        tetrachoric_correlation = get_tetrachoric_correlation(response_matrix)

        kmo_all, kmo_model = calculate_kmo(tetrachoric_correlation)
        print(kmo_model)

        # https://github.com/EducationalTestingService/factor_analyzer/blob/main/factor_analyzer/utils.py
        # line 247

    # noinspection PyMethodMayBeStatic
    def get_factor_analysis_eigen_value_list_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx"
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        response_matrix = [x.errata for x in default_checked_test_list]

        tetrachoric_correlation = get_tetrachoric_correlation(response_matrix)

        # https://thebook.io/080223/ch05/01/03/

        fa = FactorAnalyzer()
        fa.fit(tetrachoric_correlation)

        eigen_list = list(fa.get_eigenvalues()[0])
        print('1 이상 고유치의 수 : ', len([x for x in eigen_list if x > 1.0]))
        print('최대 고유치 설명량 : ', eigen_list[0]/np.sum(eigen_list))
        print('1요인과 2요인 고유치의 비 : ', eigen_list[0]/eigen_list[1])

    # noinspection PyMethodMayBeStatic
    def difficulty_distribution_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        errata_list_level_1to7 = [x.errata for x in default_checked_test_list if 1 <= int(x.level) <= 7]
        errata_list_level_8to11 = [x.errata for x in default_checked_test_list if 8 <= int(x.level) <= 11]

        estimate_1to7 = girth.twopl_mml(np.transpose(np.array(errata_list_level_1to7)))['Difficulty']
        estimate_8to11 = girth.twopl_mml(np.transpose(np.array(errata_list_level_8to11)))['Difficulty']

        unscaled_weight_list_1to7 = []
        unscaled_weight_list_8to11 = []

        for i in estimate_1to7:
            if 2.0 <= float(i):
                unscaled_weight_list_1to7.append(5)
            elif 0.5 <= float(i) < 2.0:
                unscaled_weight_list_1to7.append(4)
            elif -0.5 <= float(i) < 0.5:
                unscaled_weight_list_1to7.append(3)
            elif -2.0 <= float(i) < -0.5:
                unscaled_weight_list_1to7.append(2)
            elif float(i) < -2.0:
                unscaled_weight_list_1to7.append(1)
            else:
                raise Exception()

        for i in estimate_8to11:
            if 2.0 <= float(i):
                unscaled_weight_list_8to11.append(5)
            elif 0.5 <= float(i) < 2.0:
                unscaled_weight_list_8to11.append(4)
            elif -0.5 <= float(i) < 0.5:
                unscaled_weight_list_8to11.append(3)
            elif -2.0 <= float(i) < -0.5:
                unscaled_weight_list_8to11.append(2)
            elif float(i) < -2.0:
                unscaled_weight_list_8to11.append(1)
            else:
                raise Exception()

        scaled_weight_list_1to7 = \
            list(np.array(unscaled_weight_list_1to7) *
                 85 / np.sum(np.array(unscaled_weight_list_1to7)))

        scaled_weight_list_8to11 = \
            list(np.array(unscaled_weight_list_8to11) *
                 85 / np.sum(np.array(unscaled_weight_list_8to11)))

        plt.scatter(range(1, 171), list(scaled_weight_list_1to7) + list(scaled_weight_list_8to11))
        plt.show()

    # noinspection PyMethodMayBeStatic
    def get_ability_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        errata_list_level_1to7 = [x.errata for x in default_checked_test_list if 1 <= int(x.level) <= 7]
        errata_list_level_8to11 = [x.errata for x in default_checked_test_list if 8 <= int(x.level) <= 11]

        estimate_1to7 = girth.twopl_mml(np.transpose(np.array(errata_list_level_1to7)))
        estimate_8to11 = girth.twopl_mml(np.transpose(np.array(errata_list_level_8to11)))

        ability_1to7 = girth.ability_eap(
            np.transpose(np.array(errata_list_level_1to7)),
            estimate_1to7['Difficulty'],
            estimate_1to7['Discrimination']
        )

        ability_8to11 = girth.ability_eap(
            np.transpose(np.array(errata_list_level_8to11)),
            estimate_8to11['Difficulty'],
            estimate_8to11['Discrimination']
        )

        np.set_printoptions(threshold=np.inf, linewidth=np.inf)

        print(ability_1to7)
        print(ability_8to11)

    # noinspection PyMethodMayBeStatic
    def difficulty_spread_distribution_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r"C:\Users\dlgur\OneDrive\문서\Reports\testData.xlsx"
        sheet_test = 'Data'
        test_list = read_test_excel(path_test, sheet_test)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        errata_list_level_1to7 = [x.errata for x in default_checked_test_list if 1 <= int(x.level) <= 7]
        errata_list_level_8to11 = [x.errata for x in default_checked_test_list if 8 <= int(x.level) <= 11]

        estimate_1to7 = np.array(girth.twopl_mml(np.transpose(np.array(errata_list_level_1to7)))['Difficulty'])
        estimate_8to11 = np.array(girth.twopl_mml(np.transpose(np.array(errata_list_level_8to11)))['Difficulty'])

        weight_list_1to7 = np.zeros(85)
        weight_list_8to11 = np.zeros(85)


        for i in range(len(estimate_1to7)):
            index = estimate_1to7.argmax()
            estimate_1to7[index] = -np.inf

            if 0 <= i <= 16:
                weight_list_1to7[index] = 1.6666666635
            elif 17 <= i <= 33:
                weight_list_1to7[index] = 1.3333333325
            elif 34 <= i <= 50:
                weight_list_1to7[index] = 1.0000000015
            elif 51 <= i <= 67:
                weight_list_1to7[index] = 0.6666666705
            elif 68 <= i <= 84:
                weight_list_1to7[index] = 0.3333333310
            else:
                raise Exception()

        for i in range(len(estimate_8to11)):
            index = estimate_8to11.argmax()
            estimate_8to11[index] = -np.inf

            if 0 <= i <= 16:
                weight_list_8to11[index] = 1.6666666635
            elif 17 <= i <= 33:
                weight_list_8to11[index] = 1.3333333325
            elif 34 <= i <= 50:
                weight_list_8to11[index] = 1.0000000015
            elif 51 <= i <= 67:
                weight_list_8to11[index] = 0.6666666705
            elif 68 <= i <= 84:
                weight_list_8to11[index] = 0.3333333310
            else:
                raise Exception()

        plt.scatter(range(1, 171), list(weight_list_1to7) + list(weight_list_8to11))
        plt.show()

    # noinspection PyMethodMayBeStatic
    def write_score_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )

        """
        no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                 lc, gr, rc,
                 score, rank_total, rank_level, clazz, score_area, rank_total_area, rank_level_area
        """

        final_test_info_list = [FinalTestInfo(
            x.no, x.name, x.birth, x.phone_number, x.teacher, x.level, x.class_number, x.sector, x.school,
            x.grade, x.lc, x.gr, x.rc, x.score, None, None, None, None, None, None
        ) for x in weight_check_test_list]

    # noinspection PyMethodMayBeStatic
    def write_rank_total_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )

        """
        no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                 lc, gr, rc,
                 score, rank_total, rank_level, clazz, score_area, rank_total_area, rank_level_area
        """

        ranking_list = np.zeros(len(test_list))
        temp = copy.deepcopy(weight_check_test_list)

        for i in range(len(test_list)):
            index = int(np.array([x.score for x in temp]).argmax())
            temp[index].score = -np.inf
            ranking_list[index] = i+1

        final_test_info_list = [FinalTestInfo(
            weight_check_test_list[x].no, weight_check_test_list[x].name, weight_check_test_list[x].birth,
            weight_check_test_list[x].phone_number, weight_check_test_list[x].teacher, weight_check_test_list[x].level,
            weight_check_test_list[x].class_number, weight_check_test_list[x].sector, weight_check_test_list[x].school,
            weight_check_test_list[x].grade, weight_check_test_list[x].lc, weight_check_test_list[x].gr, weight_check_test_list[x].rc,
            weight_check_test_list[x].score, str(ranking_list[x]*100/len(test_list)) + '%', None, None, None, None, None
        ) for x in range(len(weight_check_test_list))]

        for i in final_test_info_list:
            print(i.score, ', ', i.rank_total)

    # noinspection PyMethodMayBeStatic
    def numpy_sort_test(self):
        arr = np.array([
            [1, 5, 4, 5],
            [0, 1, 9, 5]
        ])

        print(np.sort(arr, axis=1)[np.array([0, 0, 1, 0])])
        print(np.sort(arr, axis=0))

    # noinspection PyMethodMayBeStatic
    def write_rank_level_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )

        """
        no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                 lc, gr, rc,
                 score, rank_total, rank_level, clazz, score_area, rank_total_area, rank_level_area
        """

        #ranking_list = np.zeros(len(test_list))
        #temp = copy.deepcopy(weight_check_test_list)

        # 레벨별로 저장한 인덱스 리스트
        index_list_level = [[i for i in range(len(test_list)) if test_list[i].level == str(j)]
                            for j in range(1, 12)]

        np.set_printoptions(threshold=np.inf, linewidth=np.inf)

        print([test_list[x].level for x in index_list_level[0]])
        print([x for x in index_list_level[0]])

        ranking_list = np.zeros(len(test_list))
        temp = copy.deepcopy(weight_check_test_list)

        for i in range(len(test_list)):
            index = int(np.array([x.score for x in temp]).argmax())
            temp[index].score = -np.inf
            ranking_list[index] = i + 1

        final_test_info_list = [FinalTestInfo(
            weight_check_test_list[x].no, weight_check_test_list[x].name, weight_check_test_list[x].birth,
            weight_check_test_list[x].phone_number, weight_check_test_list[x].teacher, weight_check_test_list[x].level,
            weight_check_test_list[x].class_number, weight_check_test_list[x].sector, weight_check_test_list[x].school,
            weight_check_test_list[x].grade, weight_check_test_list[x].lc, weight_check_test_list[x].gr,
            weight_check_test_list[x].rc,
            weight_check_test_list[x].score, str(ranking_list[x] * 100 / len(test_list)) + '%', None, None, None, None,
            None
        ) for x in range(len(weight_check_test_list))]

        index_list_level_index_list_score = []

        for i in range(11):
            current_level_index = np.array([int(x) for x in index_list_level[i]])

            if len(current_level_index) == 0:
                index_list_level_index_list_score.append([])

            else:
                current_index_list_score = np.flip(np.argsort(
                    np.array([float(x.score) for x in weight_check_test_list])[np.array(current_level_index)]
                ))
                index_list_level_index_list_score.append(current_index_list_score)

        for i in index_list_level_index_list_score:
            print(i)

        for i in range(len(index_list_level_index_list_score)):
            ranking_count = 1

            for j in range(len(index_list_level_index_list_score[i])):
                final_test_info_list[
                    int(index_list_level[i][int(index_list_level_index_list_score[i][j])])
                ].rank_level = ranking_count
                ranking_count = ranking_count + 1

        for i in final_test_info_list:
            print(i.score, ', ', i.rank_total, ', ', i.rank_level)

    # noinspection PyMethodMayBeStatic
    def rank2stanine_test(self):
        rank = 0.1
        stanine = rank2stanine(rank)
        print(stanine)

    # noinspection PyMethodMayBeStatic
    def write_stanine_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )

        """
        no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                 lc, gr, rc,
                 score, rank_total, rank_level, stanine, score_area, rank_total_area, rank_level_area
        """

        #ranking_list = np.zeros(len(test_list))
        #temp = copy.deepcopy(weight_check_test_list)

        # 레벨별로 저장한 인덱스 리스트
        index_list_level = [[i for i in range(len(test_list)) if test_list[i].level == str(j)]
                            for j in range(1, 12)]

        np.set_printoptions(threshold=np.inf, linewidth=np.inf)

        print([test_list[x].level for x in index_list_level[0]])
        print([x for x in index_list_level[0]])

        ranking_list = np.zeros(len(test_list))
        temp = copy.deepcopy(weight_check_test_list)

        for i in range(len(test_list)):
            index = int(np.array([x.score for x in temp]).argmax())
            temp[index].score = -np.inf
            ranking_list[index] = i + 1


        ranking_list_ordered_by_level_score = np.zeros(len(test_list))
        index_list_level_index_list_score = []

        for i in range(11):
            current_level_index = np.array([int(x) for x in index_list_level[i]])

            if len(current_level_index) == 0:
                index_list_level_index_list_score.append([])
            else:
                current_index_list_score = np.flip(np.argsort(
                    np.array([float(x.score) for x in weight_check_test_list])[np.array(current_level_index)]
                ))
                index_list_level_index_list_score.append(current_index_list_score)

        for i in index_list_level_index_list_score:
            print(i)

        for i in range(len(index_list_level_index_list_score)):
            ranking_count = 1

            for j in range(len(index_list_level_index_list_score[i])):
                ranking_list_ordered_by_level_score[
                    int(index_list_level[i][int(index_list_level_index_list_score[i][j])])
                ] = float(ranking_count * 100 / len(index_list_level_index_list_score[i]))
                ranking_count = ranking_count + 1

        final_test_info_list = [FinalTestInfo(
            weight_check_test_list[x].no, weight_check_test_list[x].name, weight_check_test_list[x].birth,
            weight_check_test_list[x].phone_number, weight_check_test_list[x].teacher, weight_check_test_list[x].level,
            weight_check_test_list[x].class_number, weight_check_test_list[x].sector, weight_check_test_list[x].school,
            weight_check_test_list[x].grade, weight_check_test_list[x].lc, weight_check_test_list[x].gr,
            weight_check_test_list[x].rc,
            weight_check_test_list[x].score, str(ranking_list[x] * 100 / len(test_list)) + '%',
            str(ranking_list_ordered_by_level_score[x]) + '%',
            rank2stanine(float(ranking_list[x] * 100 / len(test_list))), None, None, None
        ) for x in range(len(weight_check_test_list))]

        for i in final_test_info_list:
            print(i.score, ', ', i.rank_total, ', ', i.rank_level, ', ', i.stanine)

    # noinspection PyMethodMayBeStatic
    def write_score_area_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )

        """
        no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                 lc, gr, rc,
                 score, rank_total, rank_level, stanine, score_area, rank_total_area, rank_level_area
        """

        #ranking_list = np.zeros(len(test_list))
        #temp = copy.deepcopy(weight_check_test_list)

        # 레벨별로 저장한 인덱스 리스트
        index_list_level = [[i for i in range(len(test_list)) if test_list[i].level == str(j)]
                            for j in range(1, 12)]

        np.set_printoptions(threshold=np.inf, linewidth=np.inf)

        print([test_list[x].level for x in index_list_level[0]])
        print([x for x in index_list_level[0]])

        ranking_list = np.zeros(len(test_list))
        temp = copy.deepcopy(weight_check_test_list)

        for i in range(len(test_list)):
            index = int(np.array([x.score for x in temp]).argmax())
            temp[index].score = -np.inf
            ranking_list[index] = i + 1


        ranking_list_ordered_by_level_score = np.zeros(len(test_list))
        index_list_level_index_list_score = []

        for i in range(11):
            current_level_index = np.array([int(x) for x in index_list_level[i]])

            if len(current_level_index) == 0:
                index_list_level_index_list_score.append([])
            else:
                current_index_list_score = np.flip(np.argsort(
                    np.array([float(x.score) for x in weight_check_test_list])[np.array(current_level_index)]
                ))
                index_list_level_index_list_score.append(current_index_list_score)

        for i in index_list_level_index_list_score:
            print(i)

        for i in range(len(index_list_level_index_list_score)):
            ranking_count = 1

            for j in range(len(index_list_level_index_list_score[i])):
                ranking_list_ordered_by_level_score[
                    int(index_list_level[i][int(index_list_level_index_list_score[i][j])])
                ] = float(ranking_count * 100 / len(index_list_level_index_list_score[i]))
                ranking_count = ranking_count + 1

        index_weighted_answer_1to7_lc_general_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'LC 일반']
        index_weighted_answer_1to7_lc_deduction_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'LC 추론']
        index_weighted_answer_1to7_gr_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].domain == 'GR']
        index_weighted_answer_1to7_rc_sat_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'RC 수능']
        index_weighted_answer_1to7_rc_general_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'RC 일반']
        index_weighted_answer_1to7_rc_deduction_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'RC 추론']

        index_dict_1to7 = {'LC 일반':index_weighted_answer_1to7_lc_general_list,
                           'LC 추론':index_weighted_answer_1to7_lc_deduction_list,
                           'GR':index_weighted_answer_1to7_gr_list,
                           'RC 수능':index_weighted_answer_1to7_rc_sat_list,
                           'RC 일반':index_weighted_answer_1to7_rc_general_list,
                           'RC 추론':index_weighted_answer_1to7_rc_deduction_list}

        index_weighted_answer_8to11_lc_general_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'LC 일반']
        index_weighted_answer_8to11_lc_deduction_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'LC 추론']
        index_weighted_answer_8to11_gr_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].domain == 'GR']
        index_weighted_answer_8to11_rc_sat_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'RC 수능']
        index_weighted_answer_8to11_rc_general_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'RC 일반']
        index_weighted_answer_8to11_rc_deduction_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'RC 추론']

        index_dict_8to11 = {'LC 일반': index_weighted_answer_8to11_lc_general_list,
                           'LC 추론': index_weighted_answer_8to11_lc_deduction_list,
                           'GR': index_weighted_answer_8to11_gr_list,
                           'RC 수능': index_weighted_answer_8to11_rc_sat_list,
                           'RC 일반': index_weighted_answer_8to11_rc_general_list,
                           'RC 추론': index_weighted_answer_8to11_rc_deduction_list}

        score_area_dict_list = []

        for i in range(len(default_checked_test_list)):
            current_score_area_dict = {
                'LC 일반': 0.0,
                'LC 추론': 0.0,
                'GR': 0.0,
                'RC 일반': 0.0,
                'RC 추론': 0.0,
                'RC 수능': 0.0
            }

            if 1 <= int(default_checked_test_list[i].level) <= 7:
                current_weight_list = [x.point for x in weighted_answer_1to7_list]
                current_index_dict = index_dict_1to7
            elif 8 <= int(default_checked_test_list[i].level) <= 11:
                current_weight_list = [x.point for x in weighted_answer_8to11_list]
                current_index_dict = index_dict_8to11
            else:
                raise Exception()

            current_score_lc_general = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['LC 일반'])],
                np.array(current_weight_list)[np.array(current_index_dict['LC 일반'])]
            )


            current_score_lc_deduction = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['LC 추론'])],
                np.array(current_weight_list)[np.array(current_index_dict['LC 추론'])]
            )

            current_score_gr = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['GR'])],
                np.array(current_weight_list)[np.array(current_index_dict['GR'])]
            )

            current_score_rc_general = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 일반'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 일반'])]
            )

            current_score_rc_deduction = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 추론'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 추론'])]
            )

            current_score_rc_sat = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 수능'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 수능'])]
            )

            current_score_area_dict['LC 일반'] = current_score_lc_general
            current_score_area_dict['LC 추론'] = current_score_lc_deduction
            current_score_area_dict['GR'] = current_score_gr
            current_score_area_dict['RC 일반'] = current_score_rc_general
            current_score_area_dict['RC 추론'] = current_score_rc_deduction
            current_score_area_dict['RC 수능'] = current_score_rc_sat

            score_area_dict_list.append(current_score_area_dict)

        final_test_info_list = [FinalTestInfo(
            weight_check_test_list[x].no, weight_check_test_list[x].name, weight_check_test_list[x].birth,
            weight_check_test_list[x].phone_number, weight_check_test_list[x].teacher, weight_check_test_list[x].level,
            weight_check_test_list[x].class_number, weight_check_test_list[x].sector, weight_check_test_list[x].school,
            weight_check_test_list[x].grade, weight_check_test_list[x].lc, weight_check_test_list[x].gr,
            weight_check_test_list[x].rc,
            weight_check_test_list[x].score, str(ranking_list[x] * 100 / len(test_list)) + '%',
            str(ranking_list_ordered_by_level_score[x]) + '%',
            rank2stanine(float(ranking_list[x] * 100 / len(test_list))), score_area_dict_list[x], None, None
        ) for x in range(len(weight_check_test_list))]

        for i in final_test_info_list:
            print(i.score, ', ', i.rank_total, ', ', i.rank_level, ', ', i.stanine, ', ', i.score_area['LC 일반'])

    # noinspection PyMethodMayBeStatic
    def write_rank_total_area_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )

        """
        no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                 lc, gr, rc,
                 score, rank_total, rank_level, stanine, score_area, rank_total_area, rank_level_area
        """

        #ranking_list = np.zeros(len(test_list))
        #temp = copy.deepcopy(weight_check_test_list)

        # 레벨별로 저장한 인덱스 리스트
        index_list_level = [[i for i in range(len(test_list)) if test_list[i].level == str(j)]
                            for j in range(1, 12)]

        np.set_printoptions(threshold=np.inf, linewidth=np.inf)

        print([test_list[x].level for x in index_list_level[0]])
        print([x for x in index_list_level[0]])

        # ranking_list = np.zeros(len(test_list))
        temp = copy.deepcopy(weight_check_test_list)
        ranking_list = score2rank([x.score for x in temp])

        """
        for i in range(len(test_list)):
            index = int(np.array([x.score for x in temp]).argmax())
            temp[index].score = -np.inf
            ranking_list[index] = i + 1
        """


        ranking_list_ordered_by_level_score = np.zeros(len(test_list))
        index_list_level_index_list_score = []

        for i in range(11):
            current_level_index = np.array([int(x) for x in index_list_level[i]])

            if len(current_level_index) == 0:
                index_list_level_index_list_score.append([])
            else:
                """
                current_index_list_score = np.flip(np.argsort(
                    np.array([float(x.score) for x in weight_check_test_list])[np.array(current_level_index)]
                ))
                index_list_level_index_list_score.append(current_index_list_score)
                """
                current_level_index_score = score2rank(
                    np.array([float(x.score) for x in weight_check_test_list])[np.array(current_level_index)]
                )
                index_list_level_index_list_score.append(current_level_index_score)

        for i in index_list_level_index_list_score:
            print(i)

        """
        for i in range(len(index_list_level_index_list_score)):
           
            ranking_count = 1

            for j in range(len(index_list_level_index_list_score[i])):
                ranking_list_ordered_by_level_score[
                    int(index_list_level[i][int(index_list_level_index_list_score[i][j])])
                ] = float(ranking_count * 100 / len(index_list_level_index_list_score[i]))
                ranking_count = ranking_count + 1
            
            for j in range(len(index_list_level_index_list_score[i])):
                ranking_list_ordered_by_level_score[
                    
                ] = float(ranking_count * 100 / len(index_list_level_index_list_score[i]))
                ranking_count = ranking_count + 1
            
        """

        for i in range(11):
            for j in range(len(index_list_level[i])):
                ranking_list_ordered_by_level_score[index_list_level[i][j]] = \
                    index_list_level_index_list_score[i][j] * 100 / len(index_list_level_index_list_score[i])

        index_weighted_answer_1to7_lc_general_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'LC 일반']
        index_weighted_answer_1to7_lc_deduction_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'LC 추론']
        index_weighted_answer_1to7_gr_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].domain == 'GR']
        index_weighted_answer_1to7_rc_sat_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'RC 수능']
        index_weighted_answer_1to7_rc_general_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'RC 일반']
        index_weighted_answer_1to7_rc_deduction_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if weighted_answer_1to7_list[i].category == 'RC 추론']

        index_dict_1to7 = {'LC 일반':index_weighted_answer_1to7_lc_general_list,
                           'LC 추론':index_weighted_answer_1to7_lc_deduction_list,
                           'GR':index_weighted_answer_1to7_gr_list,
                           'RC 수능':index_weighted_answer_1to7_rc_sat_list,
                           'RC 일반':index_weighted_answer_1to7_rc_general_list,
                           'RC 추론':index_weighted_answer_1to7_rc_deduction_list}

        index_weighted_answer_8to11_lc_general_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'LC 일반']
        index_weighted_answer_8to11_lc_deduction_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'LC 추론']
        index_weighted_answer_8to11_gr_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].domain == 'GR']
        index_weighted_answer_8to11_rc_sat_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'RC 수능']
        index_weighted_answer_8to11_rc_general_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'RC 일반']
        index_weighted_answer_8to11_rc_deduction_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if weighted_answer_8to11_list[i].category == 'RC 추론']

        index_dict_8to11 = {'LC 일반': index_weighted_answer_8to11_lc_general_list,
                           'LC 추론': index_weighted_answer_8to11_lc_deduction_list,
                           'GR': index_weighted_answer_8to11_gr_list,
                           'RC 수능': index_weighted_answer_8to11_rc_sat_list,
                           'RC 일반': index_weighted_answer_8to11_rc_general_list,
                           'RC 추론': index_weighted_answer_8to11_rc_deduction_list}

        score_area_dict_list = []

        for i in range(len(default_checked_test_list)):
            current_score_area_dict = {
                'LC 일반': 0.0,
                'LC 추론': 0.0,
                'GR': 0.0,
                'RC 일반': 0.0,
                'RC 추론': 0.0,
                'RC 수능': 0.0
            }

            if 1 <= int(default_checked_test_list[i].level) <= 7:
                current_weight_list = [x.point for x in weighted_answer_1to7_list]
                current_index_dict = index_dict_1to7
            elif 8 <= int(default_checked_test_list[i].level) <= 11:
                current_weight_list = [x.point for x in weighted_answer_8to11_list]
                current_index_dict = index_dict_8to11
            else:
                raise Exception()

            current_score_lc_general = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['LC 일반'])],
                np.array(current_weight_list)[np.array(current_index_dict['LC 일반'])]
            )


            current_score_lc_deduction = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['LC 추론'])],
                np.array(current_weight_list)[np.array(current_index_dict['LC 추론'])]
            )

            current_score_gr = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['GR'])],
                np.array(current_weight_list)[np.array(current_index_dict['GR'])]
            )

            current_score_rc_general = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 일반'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 일반'])]
            )

            current_score_rc_deduction = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 추론'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 추론'])]
            )

            current_score_rc_sat = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 수능'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 수능'])]
            )

            current_score_area_dict['LC 일반'] = current_score_lc_general
            current_score_area_dict['LC 추론'] = current_score_lc_deduction
            current_score_area_dict['GR'] = current_score_gr
            current_score_area_dict['RC 일반'] = current_score_rc_general
            current_score_area_dict['RC 추론'] = current_score_rc_deduction
            current_score_area_dict['RC 수능'] = current_score_rc_sat

            score_area_dict_list.append(current_score_area_dict)

        rank_total_area_list = []

        rank_total_area_lc_general = score2rank([x['LC 일반'] for x in score_area_dict_list])
        rank_total_area_lc_general = np.array(rank_total_area_lc_general) * 100 / len(rank_total_area_lc_general)

        rank_total_area_lc_deduction = score2rank([x['LC 추론'] for x in score_area_dict_list])
        rank_total_area_lc_deduction = np.array(rank_total_area_lc_deduction) * 100 / len(rank_total_area_lc_deduction)

        rank_total_area_gr = score2rank([x['GR'] for x in score_area_dict_list])
        rank_total_area_gr = np.array(rank_total_area_gr) * 100 / len(rank_total_area_gr)

        rank_total_area_rc_general = score2rank([x['RC 일반'] for x in score_area_dict_list])
        rank_total_area_rc_general = np.array(rank_total_area_rc_general) * 100 / len(rank_total_area_rc_general)

        rank_total_area_rc_deduction = score2rank([x['RC 추론'] for x in score_area_dict_list])
        rank_total_area_rc_deduction = np.array(rank_total_area_rc_deduction) * 100 / len(rank_total_area_rc_deduction)

        rank_total_area_rc_sat = score2rank([x['RC 수능'] for x in score_area_dict_list])
        rank_total_area_rc_sat = np.array(rank_total_area_rc_sat) * 100 / len(rank_total_area_rc_sat)

        for i in range(len(score_area_dict_list)):
            current_rank_area_dict = {'LC 일반': str(rank_total_area_lc_general[i])+'%', 'LC 추론': str(rank_total_area_lc_deduction[i])+'%',
                                      'GR': str(rank_total_area_gr[i])+'%', 'RC 일반': str(rank_total_area_rc_general[i])+'%',
                                      'RC 추론': str(rank_total_area_rc_deduction[i])+'%', 'RC 수능': str(rank_total_area_rc_sat[i])+'%'}

            rank_total_area_list.append(current_rank_area_dict)

        final_test_info_list = [FinalTestInfo(
            weight_check_test_list[x].no, weight_check_test_list[x].name, weight_check_test_list[x].birth,
            weight_check_test_list[x].phone_number, weight_check_test_list[x].teacher, weight_check_test_list[x].level,
            weight_check_test_list[x].class_number, weight_check_test_list[x].sector, weight_check_test_list[x].school,
            weight_check_test_list[x].grade, weight_check_test_list[x].lc, weight_check_test_list[x].gr,
            weight_check_test_list[x].rc,
            weight_check_test_list[x].score, str(ranking_list[x] * 100 / len(test_list)) + '%',
            str(ranking_list_ordered_by_level_score[x]) + '%',
            rank2stanine(float(ranking_list[x] * 100 / len(test_list))), score_area_dict_list[x], rank_total_area_list[x], None
        ) for x in range(len(weight_check_test_list))]


        for i in final_test_info_list:
            print(i.score, ', ', i.rank_total, ', ', i.rank_level, ', ', i.stanine, ', ', i.rank_total_area['LC 일반'])

    # noinspection PyMethodMayBeStatic
    def score2rank_test(self):
        score_list = [4.0, 4.0, 4.0, 3.0, 3.0, 9.0, 2.0, 2.0]
        rank = score2rank(score_list)
        print(rank)

    # noinspection PyMethodMayBeStatic
    def write_rank_level_area_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )

        """
        no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                 lc, gr, rc,
                 score, rank_total, rank_level, stanine, score_area, rank_total_area, rank_level_area
        """

        # ranking_list = np.zeros(len(test_list))
        # temp = copy.deepcopy(weight_check_test_list)

        # 레벨별로 저장한 인덱스 리스트
        """
        index_list_level = [[i for i in range(len(test_list)) if test_list[i].level == str(j)]
                            for j in range(1, 12)]
        """

        np.set_printoptions(threshold=np.inf, linewidth=np.inf)




        # ranking_list = np.zeros(len(test_list))
        ranking_list = score2rank([x.score for x in weight_check_test_list])

        ranking_list_ordered_by_level_score = np.array(get_score_list_group_level(
            test_list, [x.score for x in weight_check_test_list]
        ))

        index_weighted_answer_1to7_lc_general_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                      weighted_answer_1to7_list[i].category == 'LC 일반']
        index_weighted_answer_1to7_lc_deduction_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                        weighted_answer_1to7_list[i].category == 'LC 추론']
        index_weighted_answer_1to7_gr_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                              weighted_answer_1to7_list[i].domain == 'GR']
        index_weighted_answer_1to7_rc_sat_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                  weighted_answer_1to7_list[i].category == 'RC 수능']
        index_weighted_answer_1to7_rc_general_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                      weighted_answer_1to7_list[i].category == 'RC 일반']
        index_weighted_answer_1to7_rc_deduction_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                        weighted_answer_1to7_list[i].category == 'RC 추론']

        index_dict_1to7 = {'LC 일반': index_weighted_answer_1to7_lc_general_list,
                           'LC 추론': index_weighted_answer_1to7_lc_deduction_list,
                           'GR': index_weighted_answer_1to7_gr_list,
                           'RC 수능': index_weighted_answer_1to7_rc_sat_list,
                           'RC 일반': index_weighted_answer_1to7_rc_general_list,
                           'RC 추론': index_weighted_answer_1to7_rc_deduction_list}

        index_weighted_answer_8to11_lc_general_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                       weighted_answer_8to11_list[i].category == 'LC 일반']
        index_weighted_answer_8to11_lc_deduction_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                         weighted_answer_8to11_list[i].category == 'LC 추론']
        index_weighted_answer_8to11_gr_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                               weighted_answer_8to11_list[i].domain == 'GR']
        index_weighted_answer_8to11_rc_sat_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                   weighted_answer_8to11_list[i].category == 'RC 수능']
        index_weighted_answer_8to11_rc_general_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                       weighted_answer_8to11_list[i].category == 'RC 일반']
        index_weighted_answer_8to11_rc_deduction_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                         weighted_answer_8to11_list[i].category == 'RC 추론']

        index_dict_8to11 = {'LC 일반': index_weighted_answer_8to11_lc_general_list,
                            'LC 추론': index_weighted_answer_8to11_lc_deduction_list,
                            'GR': index_weighted_answer_8to11_gr_list,
                            'RC 수능': index_weighted_answer_8to11_rc_sat_list,
                            'RC 일반': index_weighted_answer_8to11_rc_general_list,
                            'RC 추론': index_weighted_answer_8to11_rc_deduction_list}

        score_area_dict_list = []
        for i in range(len(default_checked_test_list)):
            current_score_area_dict = {
                'LC 일반': 0.0,
                'LC 추론': 0.0,
                'GR': 0.0,
                'RC 일반': 0.0,
                'RC 추론': 0.0,
                'RC 수능': 0.0
            }

            if 1 <= int(default_checked_test_list[i].level) <= 7:
                current_weight_list = [x.point for x in weighted_answer_1to7_list]
                current_index_dict = index_dict_1to7
            elif 8 <= int(default_checked_test_list[i].level) <= 11:
                current_weight_list = [x.point for x in weighted_answer_8to11_list]
                current_index_dict = index_dict_8to11
            else:
                raise Exception()

            current_score_lc_general = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['LC 일반'])],
                np.array(current_weight_list)[np.array(current_index_dict['LC 일반'])]
            )

            current_score_lc_deduction = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['LC 추론'])],
                np.array(current_weight_list)[np.array(current_index_dict['LC 추론'])]
            )

            current_score_gr = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['GR'])],
                np.array(current_weight_list)[np.array(current_index_dict['GR'])]
            )

            current_score_rc_general = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 일반'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 일반'])]
            )

            current_score_rc_deduction = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 추론'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 추론'])]
            )

            current_score_rc_sat = np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 수능'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 수능'])]
            )

            current_score_area_dict['LC 일반'] = current_score_lc_general
            current_score_area_dict['LC 추론'] = current_score_lc_deduction
            current_score_area_dict['GR'] = current_score_gr
            current_score_area_dict['RC 일반'] = current_score_rc_general
            current_score_area_dict['RC 추론'] = current_score_rc_deduction
            current_score_area_dict['RC 수능'] = current_score_rc_sat

            score_area_dict_list.append(current_score_area_dict)

        rank_level_area_dict = {
            'LC 일반': get_score_list_group_level(test_list, [x['LC 일반'] for x in score_area_dict_list]),
            'LC 추론': get_score_list_group_level(test_list, [x['LC 추론'] for x in score_area_dict_list]),
            'GR': get_score_list_group_level(test_list, [x['GR'] for x in score_area_dict_list]),
            'RC 일반': get_score_list_group_level(test_list, [x['RC 일반'] for x in score_area_dict_list]),
            'RC 추론': get_score_list_group_level(test_list, [x['RC 추론'] for x in score_area_dict_list]),
            'RC 수능': get_score_list_group_level(test_list, [x['RC 수능'] for x in score_area_dict_list]),
        }

        rank_level_area_dict_list = []

        for i in range(len(score_area_dict_list)):
            rank_level_area_dict_list.append(
                {
                    'LC 일반': rank_level_area_dict['LC 일반'][i],
                    'LC 추론': rank_level_area_dict['LC 추론'][i],
                    'GR': rank_level_area_dict['GR'][i],
                    'RC 일반': rank_level_area_dict['RC 일반'][i],
                    'RC 추론': rank_level_area_dict['RC 추론'][i],
                    'RC 수능': rank_level_area_dict['RC 수능'][i],
                }
            )

        rank_total_area_list = []

        rank_total_area_lc_general = score2rank([x['LC 일반'] for x in score_area_dict_list])
        rank_total_area_lc_general = np.array(rank_total_area_lc_general) * 100 / len(rank_total_area_lc_general)

        rank_total_area_lc_deduction = score2rank([x['LC 추론'] for x in score_area_dict_list])
        rank_total_area_lc_deduction = np.array(rank_total_area_lc_deduction) * 100 / len(rank_total_area_lc_deduction)

        rank_total_area_gr = score2rank([x['GR'] for x in score_area_dict_list])
        rank_total_area_gr = np.array(rank_total_area_gr) * 100 / len(rank_total_area_gr)

        rank_total_area_rc_general = score2rank([x['RC 일반'] for x in score_area_dict_list])
        rank_total_area_rc_general = np.array(rank_total_area_rc_general) * 100 / len(rank_total_area_rc_general)

        rank_total_area_rc_deduction = score2rank([x['RC 추론'] for x in score_area_dict_list])
        rank_total_area_rc_deduction = np.array(rank_total_area_rc_deduction) * 100 / len(rank_total_area_rc_deduction)

        rank_total_area_rc_sat = score2rank([x['RC 수능'] for x in score_area_dict_list])
        rank_total_area_rc_sat = np.array(rank_total_area_rc_sat) * 100 / len(rank_total_area_rc_sat)

        for i in range(len(score_area_dict_list)):
            current_rank_area_dict = {'LC 일반': str(rank_total_area_lc_general[i]) + '%',
                                      'LC 추론': str(rank_total_area_lc_deduction[i]) + '%',
                                      'GR': str(rank_total_area_gr[i]) + '%',
                                      'RC 일반': str(rank_total_area_rc_general[i]) + '%',
                                      'RC 추론': str(rank_total_area_rc_deduction[i]) + '%',
                                      'RC 수능': str(rank_total_area_rc_sat[i]) + '%'}

            rank_total_area_list.append(current_rank_area_dict)

        final_test_info_list = [FinalTestInfo(
            weight_check_test_list[x].no, weight_check_test_list[x].name, weight_check_test_list[x].birth,
            weight_check_test_list[x].phone_number, weight_check_test_list[x].teacher, weight_check_test_list[x].level,
            weight_check_test_list[x].class_number, weight_check_test_list[x].sector, weight_check_test_list[x].school,
            weight_check_test_list[x].grade, weight_check_test_list[x].lc, weight_check_test_list[x].gr,
            weight_check_test_list[x].rc,
            weight_check_test_list[x].score, str(ranking_list[x] * 100 / len(test_list)) + '%',
            str(ranking_list_ordered_by_level_score[x]) + '%',
            rank2stanine(float(ranking_list[x] * 100 / len(test_list))), score_area_dict_list[x],
            rank_total_area_list[x], rank_level_area_dict_list[x]
        ) for x in range(len(weight_check_test_list))]

        for i in final_test_info_list:
            print(i.score, ', ', i.rank_total, ', ', i.rank_level, ', ', i.stanine, ', ', i.rank_total_area['LC 일반'])

    # noinspection PyMethodMayBeStatic
    def save_final_list_test(self):
        path_answer_1to7 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer1to7.xlsx"
        sheet_answer_1to7 = '정답'
        answer_1to7_list = read_answer_excel(path_answer_1to7, sheet_answer_1to7)

        path_answer_8to11 = r"C:\Users\dlgur\OneDrive\문서\Reports\answer8to11.xlsx"
        sheet_answer_8to11 = '정답'
        answer_8to11_list = read_answer_excel(path_answer_8to11, sheet_answer_8to11)

        path_test = r'C:\Users\dlgur\OneDrive\문서\Reports\pseudoTestData.xlsx'
        sheet_test = 'Data'
        test_list = read_pseudo_test_excel(path_test, sheet_test)

        weighted_answer_1to7_list, weighted_answer_8to11_list = \
            get_weighted_answer_list(test_list, answer_1to7_list, answer_8to11_list)

        default_checked_test_list = get_default_checked_test_list(
            test_list, answer_1to7_list, answer_8to11_list
        )

        weight_check_test_list = get_weight_checked_test_list(
            default_checked_test_list,
            weighted_answer_1to7_list,
            weighted_answer_8to11_list
        )

        """
        no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                 lc, gr, rc,
                 score, rank_total, rank_level, stanine, score_area, rank_total_area, rank_level_area
        """

        # ranking_list = np.zeros(len(test_list))
        # temp = copy.deepcopy(weight_check_test_list)

        # 레벨별로 저장한 인덱스 리스트
        """
        index_list_level = [[i for i in range(len(test_list)) if test_list[i].level == str(j)]
                            for j in range(1, 12)]
        """

        np.set_printoptions(threshold=np.inf, linewidth=np.inf)

        ranking_list = np.zeros(len(test_list))

        index_level = [[x for x in range(len(weight_check_test_list)) if 1 <= int(weight_check_test_list[x].level) <= 7],
                       [x for x in range(len(weight_check_test_list)) if 8 <= int(weight_check_test_list[x].level) <= 11]]

        index_ordered = [score2rank([x for x in np.array([x.score for x in weight_check_test_list])[np.array(index_level[0])]]),
                         score2rank([x for x in np.array([x.score for x in weight_check_test_list])[np.array(index_level[1])]])]

        for i in range(len(index_level)):
            for j in range(len(index_level[i])):
                ranking_list[index_level[i][j]] = float(index_ordered[i][j] * 100 / len(index_level[i]))

        # l [1 2 8 8 9 7 4 3 1]
        # i [0 1 5 6 7 8], [2 3 4]
        # i [6 1 4 5 3 2], [2 3 1]
        # i  6 1 2 3 1 4 3 3 2

        ranking_list_ordered_by_level_score = np.array(get_score_list_group_level(
            test_list, [x.score for x in weight_check_test_list]
        ))

        index_weighted_answer_1to7_lc_general_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                      weighted_answer_1to7_list[i].category == 'LC 일반']
        index_weighted_answer_1to7_lc_deduction_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                        weighted_answer_1to7_list[i].category == 'LC 추론']
        index_weighted_answer_1to7_gr_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                              weighted_answer_1to7_list[i].domain == 'GR']
        index_weighted_answer_1to7_rc_sat_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                  weighted_answer_1to7_list[i].category == 'RC 수능']
        index_weighted_answer_1to7_rc_general_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                      weighted_answer_1to7_list[i].category == 'RC 일반']
        index_weighted_answer_1to7_rc_deduction_list = [int(i) for i in range(len(weighted_answer_1to7_list)) if
                                                        weighted_answer_1to7_list[i].category == 'RC 추론']

        index_dict_1to7 = {'LC 일반': index_weighted_answer_1to7_lc_general_list,
                           'LC 추론': index_weighted_answer_1to7_lc_deduction_list,
                           'GR': index_weighted_answer_1to7_gr_list,
                           'RC 수능': index_weighted_answer_1to7_rc_sat_list,
                           'RC 일반': index_weighted_answer_1to7_rc_general_list,
                           'RC 추론': index_weighted_answer_1to7_rc_deduction_list}

        index_weighted_answer_8to11_lc_general_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                       weighted_answer_8to11_list[i].category == 'LC 일반']
        index_weighted_answer_8to11_lc_deduction_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                         weighted_answer_8to11_list[i].category == 'LC 추론']
        index_weighted_answer_8to11_gr_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                               weighted_answer_8to11_list[i].domain == 'GR']
        index_weighted_answer_8to11_rc_sat_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                   weighted_answer_8to11_list[i].category == 'RC 수능']
        index_weighted_answer_8to11_rc_general_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                       weighted_answer_8to11_list[i].category == 'RC 일반']
        index_weighted_answer_8to11_rc_deduction_list = [int(i) for i in range(len(weighted_answer_8to11_list)) if
                                                         weighted_answer_8to11_list[i].category == 'RC 추론']

        index_dict_8to11 = {'LC 일반': index_weighted_answer_8to11_lc_general_list,
                            'LC 추론': index_weighted_answer_8to11_lc_deduction_list,
                            'GR': index_weighted_answer_8to11_gr_list,
                            'RC 수능': index_weighted_answer_8to11_rc_sat_list,
                            'RC 일반': index_weighted_answer_8to11_rc_general_list,
                            'RC 추론': index_weighted_answer_8to11_rc_deduction_list}

        score_area_dict_list = []
        for i in range(len(default_checked_test_list)):
            current_score_area_dict = {
                'LC 일반': 0.0,
                'LC 추론': 0.0,
                'GR': 0.0,
                'RC 일반': 0.0,
                'RC 추론': 0.0,
                'RC 수능': 0.0
            }

            if 1 <= int(default_checked_test_list[i].level) <= 7:
                current_weight_list = [x.point for x in weighted_answer_1to7_list]
                current_index_dict = index_dict_1to7
                current_score_area_min = 30.0
                current_scaler = 70.0
                current_score_area_max = {
                    'LC 일반': np.sum([float(x.point) for x in weighted_answer_1to7_list if x.category == 'LC 일반']),
                    'LC 추론': np.sum([float(x.point) for x in weighted_answer_1to7_list if x.category == 'LC 추론']),
                    'GR': np.sum([float(x.point) for x in weighted_answer_1to7_list if x.domain == 'GR']),
                    'RC 일반': np.sum([float(x.point) for x in weighted_answer_1to7_list if x.category == 'RC 일반']),
                    'RC 추론': np.sum([float(x.point) for x in weighted_answer_1to7_list if x.category == 'RC 추론']),
                    'RC 수능': np.sum([float(x.point) for x in weighted_answer_1to7_list if x.category == 'RC 수능'])
                }

            elif 8 <= int(default_checked_test_list[i].level) <= 11:
                current_weight_list = [x.point for x in weighted_answer_8to11_list]
                current_index_dict = index_dict_8to11
                current_score_area_min = 40.0
                current_scaler = 60.0
                current_score_area_max = {
                    'LC 일반': np.sum([float(x.point) for x in weighted_answer_8to11_list if x.category == 'LC 일반']),
                    'LC 추론': np.sum([float(x.point) for x in weighted_answer_8to11_list if x.category == 'LC 추론']),
                    'GR': np.sum([float(x.point) for x in weighted_answer_8to11_list if x.domain == 'GR']),
                    'RC 일반': np.sum([float(x.point) for x in weighted_answer_8to11_list if x.category == 'RC 일반']),
                    'RC 추론': np.sum([float(x.point) for x in weighted_answer_8to11_list if x.category == 'RC 추론']),
                    'RC 수능': np.sum([float(x.point) for x in weighted_answer_8to11_list if x.category == 'RC 수능'])
                }

            else:
                raise Exception()

            current_score_lc_general = np.round(current_score_area_min + (current_scaler * np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['LC 일반'])],
                np.array(current_weight_list)[np.array(current_index_dict['LC 일반'])]
            )) / current_score_area_max['LC 일반'])

            current_score_lc_deduction = np.round(current_score_area_min + (current_scaler * np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['LC 추론'])],
                np.array(current_weight_list)[np.array(current_index_dict['LC 추론'])]
            )) / current_score_area_max['LC 추론'])

            current_score_gr = np.round(current_score_area_min + (current_scaler * np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['GR'])],
                np.array(current_weight_list)[np.array(current_index_dict['GR'])]
            )) / current_score_area_max['GR'])

            current_score_rc_general = np.round(current_score_area_min + (current_scaler * np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 일반'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 일반'])]
            )) / current_score_area_max['RC 일반'])

            current_score_rc_deduction = np.round(current_score_area_min + (current_scaler * np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 추론'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 추론'])]
            )) / current_score_area_max['RC 추론'])

            current_score_rc_sat = np.round(current_score_area_min + (current_scaler * np.dot(
                np.array(default_checked_test_list[i].errata)[np.array(current_index_dict['RC 수능'])],
                np.array(current_weight_list)[np.array(current_index_dict['RC 수능'])]
            )) / current_score_area_max['RC 수능'])

            current_score_area_dict['LC 일반'] = current_score_lc_general
            current_score_area_dict['LC 추론'] = current_score_lc_deduction
            current_score_area_dict['GR'] = current_score_gr
            current_score_area_dict['RC 일반'] = current_score_rc_general
            current_score_area_dict['RC 추론'] = current_score_rc_deduction
            current_score_area_dict['RC 수능'] = current_score_rc_sat

            score_area_dict_list.append(current_score_area_dict)

        rank_level_area_dict = {
            'LC 일반': [my_ceil(y, 2) for y in get_score_list_group_level(test_list, [x['LC 일반'] for x in score_area_dict_list])],
            'LC 추론': [my_ceil(y, 2) for y in get_score_list_group_level(test_list, [x['LC 추론'] for x in score_area_dict_list])],
            'GR': [my_ceil(y, 2) for y in get_score_list_group_level(test_list, [x['GR'] for x in score_area_dict_list])],
            'RC 일반': [my_ceil(y, 2) for y in get_score_list_group_level(test_list, [x['RC 일반'] for x in score_area_dict_list])],
            'RC 추론': [my_ceil(y, 2) for y in get_score_list_group_level(test_list, [x['RC 추론'] for x in score_area_dict_list])],
            'RC 수능': [my_ceil(y, 2) for y in get_score_list_group_level(test_list, [x['RC 수능'] for x in score_area_dict_list])]
        }

        rank_level_area_dict_list = []

        for i in range(len(score_area_dict_list)):
            rank_level_area_dict_list.append(
                {
                    'LC 일반': str(rank_level_area_dict['LC 일반'][i])+'%',
                    'LC 추론': str(rank_level_area_dict['LC 추론'][i])+'%',
                    'GR': str(rank_level_area_dict['GR'][i])+'%',
                    'RC 일반': str(rank_level_area_dict['RC 일반'][i])+'%',
                    'RC 추론': str(rank_level_area_dict['RC 추론'][i])+'%',
                    'RC 수능': str(rank_level_area_dict['RC 수능'][i])+'%'
                }
            )

        rank_total_area_list = []

        """
        ranking_list = np.zeros(len(test_list))

        index_level = [[x.level for x in weight_check_test_list if 1 <= x.level <= 7],
                       [x.level for x in weight_check_test_list if 8 <= x.level <= 11]]

        index_ordered = [
            score2rank([x for x in np.array([x.score for x in weight_check_test_list])[np.array(index_level[0])]]),
            score2rank([x for x in np.array([x.score for x in weight_check_test_list])[np.array(index_level[1])]])]

        for i in range(len(index_level)):
            for j in range(len(index_level[i])):
                ranking_list[index_level[i][j]] = index_ordered[i][j]
        """

        rank_total_area_lc_general = np.zeros(len(test_list))
        index_ordered = [
            score2rank([x for x in np.array([x['LC 일반'] for x in score_area_dict_list])[np.array(index_level[0])]]),
            score2rank([x for x in np.array([x['LC 일반'] for x in score_area_dict_list])[np.array(index_level[1])]])]
        for i in range(len(index_level)):
            for j in range(len(index_level[i])):
                rank_total_area_lc_general[index_level[i][j]] = my_ceil(index_ordered[i][j] * 100 / len(index_level[i]), 2)

        rank_total_area_lc_deduction = np.zeros(len(test_list))
        index_ordered = [
            score2rank([x for x in np.array([x['LC 추론'] for x in score_area_dict_list])[np.array(index_level[0])]]),
            score2rank([x for x in np.array([x['LC 추론'] for x in score_area_dict_list])[np.array(index_level[1])]])]
        for i in range(len(index_level)):
            for j in range(len(index_level[i])):
                rank_total_area_lc_deduction[index_level[i][j]] = my_ceil(index_ordered[i][j] * 100 / len(index_level[i]), 2)

        rank_total_area_gr= np.zeros(len(test_list))
        index_ordered = [
            score2rank([x for x in np.array([x['GR'] for x in score_area_dict_list])[np.array(index_level[0])]]),
            score2rank([x for x in np.array([x['GR'] for x in score_area_dict_list])[np.array(index_level[1])]])]
        for i in range(len(index_level)):
            for j in range(len(index_level[i])):
                rank_total_area_gr[index_level[i][j]] = my_ceil(index_ordered[i][j] * 100 / len(index_level[i]), 2)

        rank_total_area_rc_general = np.zeros(len(test_list))
        index_ordered = [
            score2rank([x for x in np.array([x['RC 일반'] for x in score_area_dict_list])[np.array(index_level[0])]]),
            score2rank([x for x in np.array([x['RC 일반'] for x in score_area_dict_list])[np.array(index_level[1])]])]
        for i in range(len(index_level)):
            for j in range(len(index_level[i])):
                rank_total_area_rc_general[index_level[i][j]] = my_ceil(index_ordered[i][j] * 100 / len(index_level[i]), 2)

        rank_total_area_rc_deduction = np.zeros(len(test_list))
        index_ordered = [
            score2rank([x for x in np.array([x['RC 추론'] for x in score_area_dict_list])[np.array(index_level[0])]]),
            score2rank([x for x in np.array([x['RC 추론'] for x in score_area_dict_list])[np.array(index_level[1])]])]
        for i in range(len(index_level)):
            for j in range(len(index_level[i])):
                rank_total_area_rc_deduction[index_level[i][j]] = my_ceil(index_ordered[i][j] * 100 / len(index_level[i]), 2)

        rank_total_area_rc_sat = np.zeros(len(test_list))
        index_ordered = [
            score2rank([x for x in np.array([x['RC 수능'] for x in score_area_dict_list])[np.array(index_level[0])]]),
            score2rank([x for x in np.array([x['RC 수능'] for x in score_area_dict_list])[np.array(index_level[1])]])]
        for i in range(len(index_level)):
            for j in range(len(index_level[i])):
                rank_total_area_rc_sat[index_level[i][j]] = my_ceil(index_ordered[i][j] * 100 / len(index_level[i]), 2)

        for i in range(len(score_area_dict_list)):
            current_rank_area_dict = {'LC 일반': str(rank_total_area_lc_general[i]) + '%',
                                      'LC 추론': str(rank_total_area_lc_deduction[i]) + '%',
                                      'GR': str(rank_total_area_gr[i]) + '%',
                                      'RC 일반': str(rank_total_area_rc_general[i]) + '%',
                                      'RC 추론': str(rank_total_area_rc_deduction[i]) + '%',
                                      'RC 수능': str(rank_total_area_rc_sat[i]) + '%'}

            rank_total_area_list.append(current_rank_area_dict)

        stanine_area_list = []
        for i in range(len(score_area_dict_list)):
            current_stanine_area_dict = {'LC 일반': int(rank2stanine(float(rank_total_area_lc_general[i]))),
                                      'LC 추론': int(rank2stanine(float(rank_total_area_lc_deduction[i]))),
                                      'GR': int(rank2stanine(float(rank_total_area_gr[i]))),
                                      'RC 일반': int(rank2stanine(float(rank_total_area_rc_general[i]))),
                                      'RC 추론': int(rank2stanine(float(rank_total_area_rc_deduction[i]))),
                                      'RC 수능': int(rank2stanine(float(rank_total_area_rc_sat[i])))}
            stanine_area_list.append(current_stanine_area_dict)

## todo:final_test_info가 최종 결과 컬럼 넣은 클래스
        final_test_info_list = [FinalTestInfo(
            weight_check_test_list[x].no, weight_check_test_list[x].name, weight_check_test_list[x].birth,
            weight_check_test_list[x].phone_number, weight_check_test_list[x].teacher, weight_check_test_list[x].level,
            weight_check_test_list[x].class_number, weight_check_test_list[x].sector, weight_check_test_list[x].school,
            weight_check_test_list[x].grade, weight_check_test_list[x].lc, weight_check_test_list[x].gr,
            weight_check_test_list[x].rc,
            weight_check_test_list[x].score, str(my_ceil(ranking_list[x], 2)) + '%',
            str(my_ceil(float(ranking_list_ordered_by_level_score[x]), 2)) + '%',
            int(rank2stanine(float(ranking_list[x]))), score_area_dict_list[x],
            rank_total_area_list[x], rank_level_area_dict_list[x], stanine_area_list[x]
        ) for x in range(len(weight_check_test_list))]

        """
                        no, name, birth, phone_number, teacher, level, class_number, sector, school, grade,
                        lc, gr, rc,
                        score, rank_total, rank_level, stanine, score_area, rank_total_area, rank_level_area
        """


        """
        file_path = "./test_info.json"
        with open(file_path, 'w') as outfile:
            json.dump([{
                'no':x.no, 'name':x.name, 'birth':x.birth, 'phone_number':x.phone_number,
                'teacher':x.teacher, 'level':x.level, 'class_number':x.class_number, 'sector':x.sector,
                'school':x.school, 'grade':x.grade, 'score':x.score, 'rank_total':x.rank_total, 'rank_level':x.rank_level,
                'stanine':x.stanine, 'score_area':x.score_area, 'rank_total_area':x.rank_total_area, 'rank_level_area':x.rank_level_area
            } for x in final_test_info_list], outfile, indent=4)
        """

        write_wb = Workbook()
        write_ws = write_wb.active

        count = 1
        for x in final_test_info_list:
            """
            current_dict = {
                'no':x.no, 'name':x.name, 'birth':x.birth, 'phone_number':x.phone_number,
                'teacher':x.teacher, 'level':x.level, 'class_number':x.class_number, 'sector':x.sector,
                'school':x.school, 'grade':x.grade, 'score':x.score, 'rank_total':x.rank_total, 'rank_level':x.rank_level,
                'stanine':x.stanine, 'score_area':x.score_area, 'rank_total_area':x.rank_total_area, 'rank_level_area':x.rank_level_area
            }
            """

            write_ws.cell(count, 1, x.no)
            write_ws.cell(count, 2, x.name)
            write_ws.cell(count, 3, x.birth)
            write_ws.cell(count, 4, x.phone_number)
            write_ws.cell(count, 5, x.teacher)
            write_ws.cell(count, 6, x.level)
            write_ws.cell(count, 7, x.class_number)
            write_ws.cell(count, 8, x.sector)
            write_ws.cell(count, 9, x.school)
            write_ws.cell(count, 10, x.grade)
            write_ws.cell(count, 11, x.score)
            write_ws.cell(count, 12, x.rank_total)
            write_ws.cell(count, 13, x.rank_level)
            write_ws.cell(count, 14, x.stanine)

            write_ws.cell(count, 15, x.score_area['LC 일반'])
            write_ws.cell(count, 16, x.score_area['LC 추론'])
            write_ws.cell(count, 17, x.score_area['GR'])
            write_ws.cell(count, 18, x.score_area['RC 일반'])
            write_ws.cell(count, 19, x.score_area['RC 추론'])
            write_ws.cell(count, 20, x.score_area['RC 수능'])

            write_ws.cell(count, 21, x.rank_total_area['LC 일반'])
            write_ws.cell(count, 22, x.rank_total_area['LC 추론'])
            write_ws.cell(count, 23, x.rank_total_area['GR'])
            write_ws.cell(count, 24, x.rank_total_area['RC 일반'])
            write_ws.cell(count, 25, x.rank_total_area['RC 추론'])
            write_ws.cell(count, 26, x.rank_total_area['RC 수능'])

            write_ws.cell(count, 27, x.rank_level_area['LC 일반'])
            write_ws.cell(count, 28, x.rank_level_area['LC 추론'])
            write_ws.cell(count, 29, x.rank_level_area['GR'])
            write_ws.cell(count, 30, x.rank_level_area['RC 일반'])
            write_ws.cell(count, 31, x.rank_level_area['RC 추론'])
            write_ws.cell(count, 32, x.rank_level_area['RC 수능'])

            write_ws.cell(count, 33, x.stanine_area['LC 일반'])
            write_ws.cell(count, 34, x.stanine_area['LC 추론'])
            write_ws.cell(count, 35, x.stanine_area['GR'])
            write_ws.cell(count, 36, x.stanine_area['RC 일반'])
            write_ws.cell(count, 37, x.stanine_area['RC 추론'])
            write_ws.cell(count, 38, x.stanine_area['RC 수능'])


            count = count + 1

        write_wb.save('/Users/dlgur/OneDrive/문서/Reports/결과파일12.xlsx')

    # noinspection PyMethodMayBeStatic
    def my_ceil_test(self):
        num = 3.3434564
        digit = 3

        num = np.ceil(num * (10 ** digit)) / (10 ** digit)
        print(num)

    # noinspection PyMethodMayBeStatic
    def write_json_test(self):
        file_path = "./sample.json"

        data = {'posts': []}
        data['posts'].append({
            "title": "How to get stroage size",
            "url": "https://codechacha.com/ko/get-free-and-total-size-of-volumes-in-android/",
            "draft": "false"
        })
        data['posts'].append({
            "title": "Android Q, Scoped Storage",
            "url": "https://codechacha.com/ko/android-q-scoped-storage/",
            "draft": "false"
        })
        print(data)

        with open(file_path, 'w') as outfile:
            json.dump(data, outfile, indent=4)


# 정답 데이터 읽기
# 시험 데이터 읽기
# 정오표 만들기
# 점수 매기기
# lc, gr, rc 나눠서 irt 적용(일차원성 만족시켜야 함)

# 1차원성 검증
# 행(검사자) 열(문항) 행렬 [[1 1 0][0 0 1][0 0 0]]
# pandas 로 상관 행렬 계산(문항 간의 상관관계가 행렬 성분)
# 사이킷런 pca 설명량 측정(상관 행렬 넣기)

# 능력 추정, 2모수 능력모수 추정 후 문항정보함수 측정
# 아니면 mirt(multidimensional) 모델 쓴다
# 문항정보함수 출력
# 2모수가 가장 안정적인듯 하다.(2모수의 난이도)
# 최종 채점된 데이터 -> 시험 데이터에서

# 문항 정보 함수의 합 -> 문항의 중요도
# 이것을 문항 가중치로 설정?

# 영역별 점수
# 영역별 전체 누백
# 영역별 레벨별 누백

if __name__ == '__main__':
    UnitTest().save_final_list_test()
