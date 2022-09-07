import logging
import sqlite3
import numpy as np

from openpyxl import Workbook
from openpyxl import load_workbook

class TestRepositoryTest:
    # noinspection PyMethodMayBeStatic
    def read_excel_result_11_test(self):
        test_data = TestRepository().read_excel_result_11()

        count = 1
        for i in test_data:
            print(i.items())
            count = count + 1
            if count == 10:
                break

    # noinspection PyMethodMayBeStatic
    def preprocessing_excel_result_11_test(self):
        row_begin = 3
        row_end = 4365
        column_begin = 1
        column_end = 38

        test_data = TestRepository().read_excel_result_11()
        preprocessed_test_data = TestRepository().preprocessing_excel_result_11(test_data)

        count = 1
        for i in preprocessed_test_data:
            print(i.items())
            count = count + 1
            if count == 10:
                break

    # noinspection PyMethodMayBeStatic
    def create_table_result_11_test(self):
        TestRepository().create_table_result_11('test.db')

    # noinspection PyMethodMayBeStatic
    def save_result_11_test(self):
        testRepository = TestRepository()
        test_data = testRepository.preprocessing_excel_result_11(testRepository.read_excel_result_11())

    # noinspection PyMethodMayBeStatic
    def dictionary_list_to_tuple_value_list_test(self):
        test_data = TestRepository().preprocessing_excel_result_11(
            TestRepository().read_excel_result_11()
        )

        tuple_value_list = TestRepository().dictionary_list_to_tuple_value_list(
            test_data
        )

        for i in tuple_value_list[:10]:
            print(i)

    # noinspection PyMethodMayBeStatic
    def save_result_11_db_test(self):
        testRepository = TestRepository()
        testRepository.save_result_11_db('test.db')

    # noinspection PyMethodMayBeStatic
    def drop_table_test_data_test(self):
        try:
            conn = sqlite3.connect('test.db')
            sql = 'drop table testData'
            conn.execute(sql)
            conn.commit()
            conn.close()

            print("테이블 삭제 성공")

        except (Exception, ) as err:
            print("테이블 삭제 실패")

    # noinspection PyMethodMayBeStatic
    def clear_result_11_db_test(self):
        try:
            conn = sqlite3.connect('test.db')
            sql = 'delete from testData'
            conn.execute(sql)
            conn.commit()
            conn.close()

            print("테이블 데이터 초기화 성공")

        except (Exception, ) as err:
            print("테이블 데이터 초기화 실패")

    # noinspection PyMethodMayBeStatic
    def find_area_mean_by_grade_and_level_test(self):
        print(TestRepository().find_area_mean_by_grade_and_level('test.db', 3, '3'))
        print(TestRepository().find_area_mean_by_grade_and_level('test.db', 7, '3'))
        print(TestRepository().find_area_mean_by_grade_and_level('test.db', 3, '9'))

class TestRepository:
    # noinspection PyMethodMayBeStatic
    def read_excel_result_11(self):
        row_begin = 3
        row_end = 4365
        column_begin = 1
        column_end = 38

        path = r'C:\Users\dlgur\OneDrive\문서\Reports\결과파일12.xlsx'
        sheet = 'Sheet'

        wb = load_workbook(path)
        ws = wb[sheet]

        test_data = []

        for i in range(row_begin, row_end + 1):
            test_data.append({
                'no': ws.cell(i, 1).value,
                'name': ws.cell(i, 2).value,
                'birth': ws.cell(i, 3).value,
                'phone_number': ws.cell(i, 4).value,
                'teacher_code': ws.cell(i, 5).value,
                'level': ws.cell(i, 6).value,
                'class_number': ws.cell(i, 7).value,
                'station': ws.cell(i, 8).value,
                'school_code': ws.cell(i, 9).value,
                'grade_code': ws.cell(i, 10).value,
                'total_score': ws.cell(i, 11).value,
                'total_percentile': ws.cell(i, 12).value,
                'total_percentile_group_level': ws.cell(i, 13).value,
                'total_stanine': ws.cell(i, 14).value,
                'lc_general_score': ws.cell(i, 15).value,
                'lc_deduction_score': ws.cell(i, 16).value,
                'gr_score': ws.cell(i, 17).value,
                'rc_general_score': ws.cell(i, 18).value,
                'rc_deduction_score': ws.cell(i, 19).value,
                'rc_sat_score': ws.cell(i, 20).value,
                'lc_general_percentile': ws.cell(i, 21).value,
                'lc_deduction_percentile': ws.cell(i, 22).value,
                'gr_percentile': ws.cell(i, 23).value,
                'rc_general_percentile': ws.cell(i, 24).value,
                'rc_deduction_percentile': ws.cell(i, 25).value,
                'rc_sat_percentile': ws.cell(i, 26).value,
                'lc_general_percentile_group_level': ws.cell(i, 27).value,
                'lc_deduction_percentile_group_level': ws.cell(i, 28).value,
                'gr_percentile_group_level': ws.cell(i, 29).value,
                'rc_general_percentile_group_level': ws.cell(i, 30).value,
                'rc_deduction_percentile_group_level': ws.cell(i, 31).value,
                'rc_sat_percentile_group_level': ws.cell(i, 32).value,
                'lc_general_stanine': ws.cell(i, 33).value,
                'lc_deduction_stanine': ws.cell(i, 34).value,
                'gr_stanine': ws.cell(i, 35).value,
                'rc_general_stanine': ws.cell(i, 36).value,
                'rc_deduction_stanine': ws.cell(i, 37).value,
                'rc_sat_stanine': ws.cell(i, 38).value
            })

        return test_data

    # noinspection PyMethodMayBeStatic
    def preprocessing_excel_result_11(self, test_data:list):
        row_begin = 3
        row_end = 4365
        column_begin = 1
        column_end = 38

        row_count = row_begin

        for row in test_data:
            column_count = column_begin

            for k, v in row.items():
                if row[k] is None:
                    raise Exception('빈 셀이 존재합니다.')
                else:
                    pass

                column_count = column_count + 1

            if column_count - 1 > column_end:
                raise Exception('파일의 열의 수가 적습니다.')
            elif column_count - 1 < column_end:
                raise Exception('파일의 열의 수가 많습니다.')
            else:
                pass

            # main preprocessing logic
            try:
                row['total_percentile'] = \
                    str(row['total_percentile']).replace('%', '')
                row['total_percentile_group_level'] = \
                    str(row['total_percentile_group_level']).replace('%', '')

                row['lc_general_percentile'] = \
                    str(row['lc_general_percentile']).replace('%', '')
                row['lc_deduction_percentile'] = \
                    str(row['lc_deduction_percentile']).replace('%', '')
                row['gr_percentile'] = \
                    str(row['gr_percentile']).replace('%', '')
                row['rc_general_percentile'] = \
                    str(row['rc_general_percentile']).replace('%', '')
                row['rc_deduction_percentile'] = \
                    str(row['rc_deduction_percentile']).replace('%', '')
                row['rc_sat_percentile'] = \
                    str(row['rc_sat_percentile']).replace('%', '')

                row['lc_general_percentile_group_level'] = \
                    str(row['lc_general_percentile_group_level']).replace('%','')
                row['lc_deduction_percentile_group_level'] = \
                    str(row['lc_deduction_percentile_group_level']).replace('%', '')
                row['gr_percentile_group_level'] = \
                    str(row['gr_percentile_group_level']).replace('%', '')
                row['rc_general_percentile_group_level'] = \
                    str(row['rc_general_percentile_group_level']).replace('%', '')
                row['rc_deduction_percentile_group_level'] = \
                    str(row['rc_deduction_percentile_group_level']).replace('%', '')
                row['rc_sat_percentile_group_level'] = \
                    str(row['rc_sat_percentile_group_level']).replace('%', '')

            except (Exception,):
                logging.error(row_count, ' 에서 전처리 예외 발생.')
            #

            row_count = row_count + 1

        if row_count - 1 > row_end:
            raise Exception('파일의 행의 수가 적습니다.')
        elif row_count - 1 < row_end:
            raise Exception('파일의 행의 수가 많습니다.')
        else:
            pass

        return test_data

    # noinspection PyMethodMayBeStatic
    def create_table_result_11(self, db_name:str):
        try:
            sql = "create table if not exists testData(" \
                "no INTEGER PRIMARY KEY, " \
                "name TEXT, " \
                "birth TEXT, " \
                "phone_number TEXT, " \
                "teacher_code TEXT, " \
                "level INTEGER, " \
                "class_number TEXT, " \
                "station TEXT, " \
                "school_code TEXT, " \
                "grade_code TEXT, " \
                "total_score INTEGER, " \
                "total_percentile REAL, " \
                "total_percentile_group_level REAL, " \
                "total_stanine INTEGER, " \
                "lc_general_score INTEGER, " \
                "lc_deduction_score INTEGER, " \
                "gr_score INTEGER, " \
                "rc_general_score INTEGER, " \
                "rc_deduction_score INTEGER, " \
                "rc_sat_score INTEGER, " \
                "lc_general_percentile REAL, " \
                "lc_deduction_percentile REAL, " \
                "gr_percentile REAL, " \
                "rc_general_percentile REAL, " \
                "rc_deduction_percentile REAL, " \
                "rc_sat_percentile REAL, " \
                "lc_general_percentile_group_level REAL, " \
                "lc_deduction_percentile_group_level REAL, " \
                "gr_percentile_group_level REAL, " \
                "rc_general_percentile_group_level REAL, " \
                "rc_deduction_percentile_group_level REAL, " \
                "rc_sat_percentile_group_level REAL, " \
                "lc_general_stanine INTEGER, " \
                "lc_deduction_stanine INTEGER, " \
                "gr_stanine INTEGER, " \
                "rc_general_stanine INTEGER, " \
                "rc_deduction_stanine INTEGER, " \
                "rc_sat_stanine INTEGER " \
                ")"
            conn = sqlite3.connect(db_name)
            conn.execute(sql)
            conn.close()
            print("테이블 생성 성공")
        except (Exception, ) as err:
            print("테이블 생성 실패")

    # noinspection PyMethodMayBeStatic
    def clear_table_result_11(self, db_name:str):
        try:
            conn = sqlite3.connect(db_name)
            sql = 'delete from testData'
            conn.execute(sql)
            conn.commit()
            conn.close()

            print("테이블 데이터 초기화 성공")

        except (Exception, ) as err:
            print("테이블 데이터 초기화 실패")

    # noinspection PyMethodMayBeStatic
    def save_result_11_db(self, db_name:str):
        testRepository = TestRepository()
        test_data_tuple_list = testRepository.dictionary_list_to_tuple_value_list(
            testRepository.preprocessing_excel_result_11(
                testRepository.read_excel_result_11()
            )
        )

        testRepository.create_table_result_11(db_name)
        testRepository.clear_table_result_11(db_name)

        try:
            conn = sqlite3.connect('test.db')
            stmt = "insert into testData (" \
                   "no, name, birth, phone_number, teacher_code, level, class_number, " \
                   "station, school_code, grade_code, total_score, total_percentile, " \
                   "total_percentile_group_level, total_stanine, " \
                   "lc_general_score, lc_deduction_score, gr_score, " \
                   "rc_general_score, rc_deduction_score, rc_sat_score, " \
                   "lc_general_percentile, lc_deduction_percentile, gr_percentile, " \
                   "rc_general_percentile, rc_deduction_percentile, rc_sat_percentile, " \
                   "lc_general_percentile_group_level, lc_deduction_percentile_group_level, gr_percentile_group_level, " \
                   "rc_general_percentile_group_level, rc_deduction_percentile_group_level, rc_sat_percentile_group_level, " \
                   "lc_general_stanine, lc_deduction_stanine, gr_stanine, " \
                   "rc_general_stanine, rc_deduction_stanine, rc_sat_stanine" \
                   ") values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, " \
                   "?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, " \
                   "?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?" \
                   ")"

            conn.executemany(stmt, test_data_tuple_list)
            conn.commit()
            conn.close()

            print("데이터 추가 성공")

        except (Exception, ) as err:
            print("데이터 추가 실패")

    # noinspection PyMethodMayBeStatic
    def dictionary_list_to_tuple_value_list(self, test_data:list):
        tuple_value_list = []

        for i in test_data:
            tuple_value_list.append(tuple(dict(i).values()))

        return tuple_value_list

    # noinspection PyMethodMayBeStatic
    def drop_table_result_11(self, db_name:str):
        try:
            conn = sqlite3.connect(db_name)
            sql = 'drop table testData'
            conn.execute(sql)
            conn.commit()
            conn.close()

            print("테이블 삭제 성공")

        except (Exception, ) as err:
            print("테이블 삭제 실패")

    # noinspection PyMethodMayBeStatic
    def find_area_mean_by_grade_and_level(self, db_name:str, level:int, grade_code:str):
        if (1 <= int(level) <= 11) and ((grade_code == '1') or (grade_code == '2') or (grade_code == '3')):
            try:
                params = (level, grade_code)
                sql = "select " \
                      "  avg(lc_general_score)" \
                      ", avg(lc_deduction_score)" \
                      ", avg(gr_score)" \
                      ", avg(rc_general_score)" \
                      ", avg(rc_deduction_score)" \
                      ", avg(rc_sat_score)" \
                      ", avg(total_score)" \
                      " from testData where level = ? and grade_code = ?"
                db = sqlite3.connect(db_name)
                cur = db.cursor()
                cur.execute(sql, params)
                dt = cur.fetchall()
                db.close()

                print("조회 성공")

                return dt[0]


            except (Exception,) as err:
                print("조회 실패")

        else:
            raise Exception('level 혹은 grade 값이 잘못되었습니다.')


class SQLiteTest:
    # noinspection PyMethodMayBeStatic
    def create_table_test(self):
        try:
            sql = "create table if not exists student(name varchar(20), age int, birth date)"
            conn = sqlite3.connect('test.db')
            conn.execute(sql)
            conn.close()
            print("테이블 생성 성공")
        except (Exception, ) as err:
            print("테이블 생성 실패")

    # noinspection PyMethodMayBeStatic
    def insert_data_test(self):
        name = '신사임당'
        age = 50
        birth = '2000-01-01'

        try:
            sql = "insert into student values(?, ?, ?)"
            conn = sqlite3.connect('test.db')
            cur = conn.cursor()
            cur.execute(sql, (name, age, birth))
            conn.commit()
            conn.close()
            print("데이터 추가 성공")

        except (Exception, ) as err:
            print("데이터 추가 실패")

    # noinspection PyMethodMayBeStatic
    def select_data_test(self):
        try:
            sql = "select * from student"
            db = sqlite3.connect("test.db")
            cur = db.cursor()
            cur.execute(sql)
            dt = cur.fetchall()
            db.close()

            for n, a, b in dt:
                print(n, a, b)

        except (Exception, ) as err:
            print("select 실패")

    # noinspection PyMethodMayBeStatic
    def tuple_insert_test(self):
        try:
            conn = sqlite3.connect('test.db')
            stmt = "insert into student (name, age, birth) values (?, ?, ?)"

            cur = conn.cursor()
            ## many rows
            vals = [('종수', '12', '2003-03-21'), ('종수', 12.0, '2003-03-22')]

            conn.executemany(stmt, vals)
            conn.commit()
            conn.close()

            print("데이터 추가 성공")

        except (Exception, ) as err:
            print("데이터 추가 실패")




if __name__ == '__main__':
    print('[특목 3학년 시험 데이터]')
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 1, '3'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 2, '3'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 3, '3'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 4, '3'))
    print('')

    print('[특목 2학년 시험 데이터]')
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 1, '2'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 2, '2'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 3, '2'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 4, '2'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 5, '2'))
    print('')

    print('[특목 1학년 시험 데이터]')
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 1, '1'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 2, '1'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 3, '1'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 4, '1'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 5, '1'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 6, '1'))
    print('')

    print('[중등 3학년 시험 데이터]')
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 8, '3'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 9, '3'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 10, '3'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 11, '3'))
    print('')

    print('[중등 2학년 시험 데이터]')
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 8, '2'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 9, '2'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 10, '2'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 11, '2'))
    print('')

    print('[중등 1학년 시험 데이터]')
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 8, '1'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 9, '1'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 10, '1'))
    print(TestRepository().find_area_mean_by_grade_and_level('test.db', 11, '1'))
    print('')